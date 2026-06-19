import streamlit as st
import sqlite3
import pandas as pd
import json
import os
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path
from io import BytesIO
import base64
import requests

# --- CẤU HÌNH TRANG WEB ---
st.set_page_config(
    page_title="Tennis Vui - Quản lý Tài chính & Giải đấu",
    page_icon="🎾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- KHỞI TẠO CÁC HẰNG SỐ ---
APP_TITLE = "Tennis vui"
DEFAULT_DB_FILE = "tennis_monthly.tennis"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "Vui" # Bạn có thể đổi mật khẩu ở đây

DEFAULT_RULES = {
    "fine_lose": 20000,
    "fine_draw": 20000,
    "fine_lose_zero": 40000,
    "point_win": 3,
    "point_draw": 1,
    "point_loss": 0,
}

EXPENSE_TYPES = ["Thuê sân", "Mua bóng", "Lượm banh", "Ăn uống", "Tiệc cuối tháng", "Chi phí khác"]
WEIGHTED_EXPENSE_TYPES = {"Thuê sân", "Mua bóng", "Lượm banh"}
BALL_PICKER_EXPENSE_TYPES = {"Mua bóng", "Lượm banh"}
FUND_ONLY_EXPENSE_TYPES = {"Tiệc cuối tháng"}
ALL_MEMBER_DEFAULT_EXPENSE_TYPES = WEIGHTED_EXPENSE_TYPES.union(FUND_ONLY_EXPENSE_TYPES)
KNOWN_EXPENSE_TYPES = WEIGHTED_EXPENSE_TYPES.union({"Ăn uống"}, FUND_ONLY_EXPENSE_TYPES)

DEFAULT_INCOME_TYPES = ["Thu quỹ", "Thu vãng lai", "Thu khác"]
AUTO_ROUNDED_MEMBER_INCOME = "Thu hội viên làm tròn"
AUTO_EXTERNAL_INCOME = "Thu người ngoài"
AUTO_INCOME_TYPES = {AUTO_ROUNDED_MEMBER_INCOME, AUTO_EXTERNAL_INCOME}
FUND_ENTITY_NAME = "Quỹ"
LEGACY_FUND_ENTITY_NAMES = {FUND_ENTITY_NAME, "Chủ sân"}

# --- CÁC HÀM HỖ TRỢ TỪ TENNIS VUI ---
def normalize_text_key(value):
    text = str(value or "").strip().lower().replace("đ", "d")
    text = "".join(
        char for char in unicodedata.normalize("NFD", text)
        if not unicodedata.combining(char)
    )
    for old, new in ((" ", "_"), ("-", "_"), ("/", "_"), (".", "")):
        text = text.replace(old, new)
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def load_json_list(value):
    try:
        data = json.loads(value or "[]")
    except Exception:
        return []
    return data if isinstance(data, list) else []


def includes_all_or_empty(value):
    participants = load_json_list(value)
    return "All" in participants or not participants


def previous_month(month):
    try:
        y, m = [int(x) for x in month.split("-")]
        if m == 1:
            return f"{y-1}-12"
        return f"{y}-{m-1:02d}"
    except Exception:
        return datetime.now().strftime("%Y-%m")


def perform_auto_daily_backup(selected_db):
    try:
        db_path = Path(selected_db)
        if not db_path.exists():
            return
        
        # Không tạo bản sao lưu cho chính file sao lưu hoặc file backup khác
        if db_path.name.startswith("daily_backup_") or db_path.name.startswith("backup_"):
            return
            
        today_str = datetime.now().strftime("%Y%m%d")
        backup_prefix = f"daily_backup_{today_str}_"
        backup_name = f"{backup_prefix}{db_path.name}"
        
        workspace_dir = db_path.parent
        if str(workspace_dir) == ".":
            workspace_dir = Path(os.getcwd())
            
        # Kiểm tra xem hôm nay đã sao lưu file này chưa
        existing_backups = [
            f for f in os.listdir(workspace_dir)
            if f.startswith(backup_prefix) and f.endswith(db_path.name)
        ]
        
        if not existing_backups:
            shutil.copy(db_path, workspace_dir / backup_name)
            
        # Chỉ giữ lại tối đa 5 bản sao lưu tự động cho file này
        all_daily_backups = sorted([
            f for f in os.listdir(workspace_dir)
            if f.startswith("daily_backup_") and f.endswith(db_path.name)
        ])
        
        while len(all_daily_backups) > 5:
            oldest_backup = all_daily_backups.pop(0)
            try:
                os.remove(workspace_dir / oldest_backup)
            except Exception:
                pass
    except Exception as e:
        print(f"[ERROR] Auto daily backup failed: {e}")


def save_db_to_github(db_path_str, github_filename):
    if "GITHUB_TOKEN" not in st.secrets or "GITHUB_REPO" not in st.secrets:
        return False, "Chưa cấu hình GITHUB_TOKEN hoặc GITHUB_REPO trong Streamlit Secrets."
        
    token = st.secrets["GITHUB_TOKEN"]
    repo = st.secrets["GITHUB_REPO"]
    branch = st.secrets.get("GITHUB_BRANCH", "main")
    
    path = github_filename
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }
    
    try:
        sha = None
        r_get = requests.get(url, headers=headers)
        if r_get.status_code == 200:
            sha = r_get.json().get("sha")
            
        with open(db_path_str, "rb") as f:
            content_bytes = f.read()
        content_base64 = base64.b64encode(content_bytes).decode("utf-8")
        
        data = {
            "message": f"Backup database: {db_path_str} via Tennis Vui Web App as {github_filename}",
            "content": content_base64,
            "branch": branch
        }
        if sha:
            data["sha"] = sha
            
        r_put = requests.put(url, headers=headers, json=data)
        if r_put.status_code in [200, 201]:
            return True, f"Đã đồng bộ thành công file `{db_path_str}` lên GitHub dưới tên `{github_filename}`!"
        else:
            return False, f"Lỗi GitHub API ({r_put.status_code}): {r_put.text}"
            
    except Exception as e:
        return False, f"Lỗi kết nối GitHub: {e}"


def list_files_on_github():
    if "GITHUB_TOKEN" not in st.secrets or "GITHUB_REPO" not in st.secrets:
        return []
    token = st.secrets["GITHUB_TOKEN"]
    repo = st.secrets["GITHUB_REPO"]
    branch = st.secrets.get("GITHUB_BRANCH", "main")
    
    url = f"https://api.github.com/repos/{repo}/contents/?ref={branch}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }
    try:
        r = requests.get(url, headers=headers)
        if r.status_code == 200:
            files = r.json()
            return [f["name"] for f in files if f["type"] == "file" and f["name"].endswith(".tennis")]
    except Exception:
        pass
    return []


def load_db_from_github(github_filename, local_filename):
    if "GITHUB_TOKEN" not in st.secrets or "GITHUB_REPO" not in st.secrets:
        return False, "Chưa cấu hình GITHUB_TOKEN hoặc GITHUB_REPO trong Streamlit Secrets."
        
    token = st.secrets["GITHUB_TOKEN"]
    repo = st.secrets["GITHUB_REPO"]
    branch = st.secrets.get("GITHUB_BRANCH", "main")
    
    path = github_filename
    url = f"https://api.github.com/repos/{repo}/contents/{path}?ref={branch}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3.raw"
    }
    
    try:
        r = requests.get(url, headers=headers)
        if r.status_code == 200:
            with open(local_filename, "wb") as f:
                f.write(r.content)
            return True, f"Đã tải thành công file `{github_filename}` từ GitHub về và lưu thành `{local_filename}`!"
        else:
            return False, f"Lỗi GitHub API ({r.status_code}): {r.text}"
            
    except Exception as e:
        return False, f"Lỗi kết nối GitHub: {e}"


# --- PHẦN XỬ LÝ DATABASE ---
class TennisDB:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.create_tables()

    def create_tables(self):
        cur = self.conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS players (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                active INTEGER NOT NULL DEFAULT 1,
                weekly_sessions INTEGER NOT NULL DEFAULT 3
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                match_date TEXT NOT NULL,
                match_no INTEGER NOT NULL DEFAULT 1,
                team_a1 TEXT NOT NULL,
                team_a2 TEXT NOT NULL,
                team_b1 TEXT NOT NULL,
                team_b2 TEXT NOT NULL,
                score_a INTEGER NOT NULL,
                score_b INTEGER NOT NULL,
                match_money INTEGER,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                match_bet INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS monthly_finance (
                month TEXT PRIMARY KEY,
                court_cost INTEGER NOT NULL DEFAULT 0,
                ball_cost INTEGER NOT NULL DEFAULT 0,
                picker_cost INTEGER NOT NULL DEFAULT 0,
                reserve_fund INTEGER NOT NULL DEFAULT 1800000,
                fund_balance INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS finance_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month TEXT NOT NULL,
                name TEXT NOT NULL,
                member_type INTEGER NOT NULL DEFAULT 3
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS finance_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month TEXT NOT NULL,
                expense_date TEXT NOT NULL,
                expense_type TEXT NOT NULL,
                amount INTEGER NOT NULL DEFAULT 0,
                participants TEXT NOT NULL DEFAULT '[]',
                payer TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT ''
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS finance_incomes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month TEXT NOT NULL,
                income_date TEXT NOT NULL,
                income_type TEXT NOT NULL DEFAULT 'Thu khác',
                amount INTEGER NOT NULL DEFAULT 0,
                collector TEXT NOT NULL DEFAULT 'Quỹ',
                note TEXT NOT NULL DEFAULT ''
            )
            """
        )
        self.conn.commit()
        self.ensure_schema_columns()
        self.ensure_default_rules()

    def table_columns(self, table_name):
        try:
            cur = self.conn.cursor()
            return [col[1] for col in cur.execute(f"PRAGMA table_info({table_name})").fetchall()]
        except Exception:
            return []

    def ensure_column(self, table_name, column_name, alter_sql):
        if column_name not in self.table_columns(table_name):
            self.conn.execute(alter_sql)

    def ensure_schema_columns(self):
        migrations = [
            ("matches", "match_no", "ALTER TABLE matches ADD COLUMN match_no INTEGER NOT NULL DEFAULT 1"),
            ("matches", "match_money", "ALTER TABLE matches ADD COLUMN match_money INTEGER"),
            ("matches", "match_bet", "ALTER TABLE matches ADD COLUMN match_bet INTEGER NOT NULL DEFAULT 0"),
            ("matches", "note", "ALTER TABLE matches ADD COLUMN note TEXT NOT NULL DEFAULT ''"),
            ("players", "weekly_sessions", "ALTER TABLE players ADD COLUMN weekly_sessions INTEGER NOT NULL DEFAULT 3"),
            ("monthly_finance", "fund_balance", "ALTER TABLE monthly_finance ADD COLUMN fund_balance INTEGER NOT NULL DEFAULT 0"),
            ("finance_expenses", "payer", "ALTER TABLE finance_expenses ADD COLUMN payer TEXT NOT NULL DEFAULT ''"),
            ("finance_expenses", "note", "ALTER TABLE finance_expenses ADD COLUMN note TEXT NOT NULL DEFAULT ''"),
            ("finance_incomes", "collector", "ALTER TABLE finance_incomes ADD COLUMN collector TEXT NOT NULL DEFAULT 'Quỹ'"),
            ("finance_incomes", "note", "ALTER TABLE finance_incomes ADD COLUMN note TEXT NOT NULL DEFAULT ''"),
        ]
        for table_name, column_name, alter_sql in migrations:
            self.ensure_column(table_name, column_name, alter_sql)
        self.conn.commit()

    def ensure_default_rules(self):
        cur = self.conn.cursor()
        for k, v in DEFAULT_RULES.items():
            cur.execute("INSERT OR IGNORE INTO settings(key, value) VALUES (?, ?)", (k, str(v)))
        self.conn.commit()

    def get_rules(self):
        cur = self.conn.cursor()
        rows = dict(cur.execute("SELECT key, value FROM settings").fetchall())
        result = {}
        for k, v in DEFAULT_RULES.items():
            try:
                result[k] = int(rows.get(k, v))
            except ValueError:
                result[k] = v
        return result

    def save_rules(self, rules_dict):
        cur = self.conn.cursor()
        for k, v in rules_dict.items():
            cur.execute("INSERT OR REPLACE INTO settings(key, value) VALUES (?, ?)", (k, str(v)))
        self.conn.commit()

    def get_player_names(self):
        cur = self.conn.cursor()
        return [r[0] for r in cur.execute("SELECT name FROM players ORDER BY name").fetchall()]

    def get_finance_actor_names(self):
        names = [FUND_ENTITY_NAME]
        fund_key = normalize_text_key(FUND_ENTITY_NAME)
        for name in self.get_player_names():
            if normalize_text_key(name) != fund_key:
                names.append(name)
        return names

    def get_fund_actor_keys(self):
        return {normalize_text_key(name) for name in LEGACY_FUND_ENTITY_NAMES}

    def get_auto_income_collector(self, income_type, month):
        key = f"auto_income_collector::{month}::{income_type}"
        try:
            row = self.conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
            value = row[0].strip() if row and row[0] else ""
        except Exception:
            value = ""

        if value in self.get_finance_actor_names():
            return value
        return FUND_ENTITY_NAME

    def get_previous_month_fund(self, month):
        prev = previous_month(month)
        cur = self.conn.cursor()
        row = cur.execute("SELECT fund_balance FROM monthly_finance WHERE month = ?", (prev,)).fetchone()
        if row and row[0] is not None:
            return int(row[0])
        return 0

    def reset_sqlite_sequence(self, table_name, count):
        try:
            self.conn.execute("UPDATE sqlite_sequence SET seq = ? WHERE name = ?", (count, table_name))
        except Exception:
            pass

    def reorder_finance_members_silent(self, month):
        cur = self.conn.cursor()
        rows = cur.execute("""
            SELECT month, name, member_type
            FROM finance_members
            ORDER BY month, LOWER(name), id
        """).fetchall()
        try:
            cur.execute("DELETE FROM finance_members")
            cur.executemany(
                "INSERT INTO finance_members(id, month, name, member_type) VALUES (?, ?, ?, ?)",
                [(idx, m, name, member_type) for idx, (m, name, member_type) in enumerate(rows, start=1)]
            )
            self.reset_sqlite_sequence("finance_members", len(rows))
            self.conn.commit()
        except Exception:
            self.conn.rollback()

    def reorder_finance_expenses_silent(self, month):
        cur = self.conn.cursor()
        rows = cur.execute("""
            SELECT expense_date, expense_type, amount, participants, payer, note
            FROM finance_expenses
            WHERE month = ?
            ORDER BY expense_date, id
        """, (month,)).fetchall()
        try:
            cur.execute("DELETE FROM finance_expenses WHERE month = ?", (month,))
            cur.executemany(
                """
                INSERT INTO finance_expenses(id, month, expense_date, expense_type, amount, participants, payer, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [(idx, month, *row) for idx, row in enumerate(rows, start=1)]
            )
            self.reset_sqlite_sequence("finance_expenses", len(rows))
            self.conn.commit()
        except Exception:
            self.conn.rollback()

    def reorder_finance_incomes_silent(self, month):
        cur = self.conn.cursor()
        rows = cur.execute("""
            SELECT income_date, income_type, amount, collector, note
            FROM finance_incomes
            WHERE month = ?
            ORDER BY income_date, id
        """, (month,)).fetchall()
        try:
            cur.execute("DELETE FROM finance_incomes WHERE month = ?", (month,))
            cur.executemany(
                """
                INSERT INTO finance_incomes(id, month, income_date, income_type, amount, collector, note)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                [(idx, month, *row) for idx, row in enumerate(rows, start=1)]
            )
            self.reset_sqlite_sequence("finance_incomes", len(rows))
            self.conn.commit()
        except Exception:
            self.conn.rollback()

    def fetch_matches(self, months):
        cur = self.conn.cursor()
        placeholders = ",".join("?" for _ in months)
        return cur.execute(
            f"""
            SELECT id, match_date, match_no, team_a1, team_a2, team_b1, team_b2, score_a, score_b, match_money, match_bet, note
            FROM matches
            WHERE substr(match_date, 1, 7) IN ({placeholders})
            ORDER BY match_date DESC, match_no DESC, id DESC
            """,
            months
        ).fetchall()

    def get_period_match_money_detail_map(self, months):
        cur = self.conn.cursor()
        placeholders = ",".join("?" for _ in months)
        rows = cur.execute(
            f"""
            SELECT id, match_date, match_no, team_a1, team_a2, team_b1, team_b2,
                   score_a, score_b, match_money, match_bet
            FROM matches
            WHERE substr(match_date, 1, 7) IN ({placeholders})
            ORDER BY match_date, match_no, id
            """,
            months
        ).fetchall()

        rules = self.get_rules()
        money_map = {}

        def get_row(name):
            name = (name or "").strip()
            if not name:
                return None
            if name not in money_map:
                money_map[name] = {"match_money": 0, "bet_money": 0, "total": 0}
            return money_map[name]

        def add_money(name, match_part, bet_part):
            row = get_row(name)
            if row is None:
                return
            row["match_money"] += int(match_part or 0)
            row["bet_money"] += int(bet_part or 0)
            row["total"] += int(match_part or 0) + int(bet_part or 0)

        for _mid, _date, _no, a1, a2, b1, b2, sa, sb, match_money, match_bet in rows:
            team_a = [x for x in [a1, a2] if x]
            team_b = [x for x in [b1, b2] if x]

            base = match_money if match_money is not None else rules["fine_lose"]
            draw = match_money if match_money is not None else rules["fine_draw"]
            zero = match_money if match_money is not None else rules["fine_lose_zero"]
            bet = int(match_bet or 0)

            if sa > sb:
                for name in team_b:
                    add_money(name, zero if sb == 0 else base, bet)
            elif sb > sa:
                for name in team_a:
                    add_money(name, zero if sa == 0 else base, bet)
            else:
                for name in team_a + team_b:
                    add_money(name, draw, bet)

        return money_map

    def validate_match_data(self, data):
        try:
            datetime.strptime(data["match_date"], "%Y-%m-%d")
        except ValueError:
            raise ValueError("Ngày phải có dạng YYYY-MM-DD")
        team_a_names = [data["team_a1"].strip(), data["team_a2"].strip()]
        team_b_names = [data["team_b1"].strip(), data["team_b2"].strip()]
        team_a_used = [n for n in team_a_names if n]
        team_b_used = [n for n in team_b_names if n]

        if not team_a_used or not team_b_used:
            raise ValueError("Vui lòng nhập ít nhất 1 người cho mỗi đội")
        if len(team_a_used) != len(team_b_used):
            raise ValueError("Trận đánh đơn nhập 1 người mỗi đội; trận đánh đôi nhập 2 người mỗi đội")

        all_names = team_a_used + team_b_used
        if len(set(n.lower() for n in all_names)) < len(all_names):
            raise ValueError("Một người không được xuất hiện 2 lần trong cùng một trận")
        if data["score_a"] < 0 or data["score_b"] < 0 or data["score_a"] > 7 or data["score_b"] > 7:
            raise ValueError("Số game nên nằm trong khoảng 0 đến 7")
        if data.get("match_money") is not None and data["match_money"] < 0:
            raise ValueError("Tiền trận không được nhỏ hơn 0")
        if data.get("match_bet", 0) < 0:
            raise ValueError("Tiền độ không được nhỏ hơn 0")
        if data["score_a"] == 0 and data["score_b"] == 0:
            raise ValueError("Tỷ số 0-0 không hợp lệ")

    def calculate_stats(self, months):
        rows = self.fetch_matches(months)
        rules = self.get_rules()
        stats = {}

        def get_player(name):
            if name not in stats:
                stats[name] = {
                    "name": name,
                    "matches": 0,
                    "wins": 0,
                    "draws": 0,
                    "losses": 0,
                    "gf": 0,
                    "ga": 0,
                    "match_money": 0,
                    "bet_money": 0,
                    "money": 0,
                }
            return stats[name]

        for _, _, _, a1, a2, b1, b2, sa, sb, match_money, match_bet, _note in rows:
            team_a = [name for name in [a1, a2] if name]
            team_b = [name for name in [b1, b2] if name]
            for name in team_a + team_b:
                get_player(name)["matches"] += 1
            for name in team_a:
                p = get_player(name)
                p["gf"] += sa
                p["ga"] += sb
            for name in team_b:
                p = get_player(name)
                p["gf"] += sb
                p["ga"] += sa

            lose_match_money = match_money if match_money is not None else rules["fine_lose"]
            draw_match_money = match_money if match_money is not None else rules["fine_draw"]
            zero_match_money = match_money if match_money is not None else rules["fine_lose_zero"]
            bet_money = int(match_bet or 0)

            def add_money(player_name, match_part):
                p = get_player(player_name)
                p["match_money"] += int(match_part or 0)
                p["bet_money"] += bet_money
                p["money"] += int(match_part or 0) + bet_money

            if sa > sb:
                for name in team_a:
                    get_player(name)["wins"] += 1
                for name in team_b:
                    p = get_player(name)
                    p["losses"] += 1
                    add_money(name, zero_match_money if sb == 0 else lose_match_money)
            elif sb > sa:
                for name in team_b:
                    get_player(name)["wins"] += 1
                for name in team_a:
                    p = get_player(name)
                    p["losses"] += 1
                    add_money(name, zero_match_money if sa == 0 else lose_match_money)
            else:
                for name in team_a + team_b:
                    p = get_player(name)
                    p["draws"] += 1
                    add_money(name, draw_match_money)
        return list(stats.values())

    def calculate_finance_rows(self, month):
        cur = self.conn.cursor()
        
        cur.execute(
            "INSERT OR IGNORE INTO monthly_finance(month, court_cost, ball_cost, picker_cost, reserve_fund, fund_balance) VALUES (?, 0, 0, 0, 1800000, 0)",
            (month,)
        )

        players = cur.execute("SELECT name, weekly_sessions FROM players ORDER BY name").fetchall()
        existing = {
            r[0].lower(): r[1]
            for r in cur.execute("SELECT name, id FROM finance_members WHERE month = ?", (month,)).fetchall()
        }

        for name, weekly in players:
            weekly = int(weekly or 3)
            if weekly not in [1, 2, 3]:
                weekly = 3
            if name.lower() in existing:
                cur.execute("UPDATE finance_members SET member_type = ? WHERE id = ?", (weekly, existing[name.lower()]))
            else:
                cur.execute("INSERT INTO finance_members(month, name, member_type) VALUES (?, ?, ?)", (month, name, weekly))
        self.conn.commit()

        self.reorder_finance_expenses_silent(month)
        self.reorder_finance_incomes_silent(month)
        self.reorder_finance_members_silent(month)

        expenses = cur.execute(
            """
            SELECT id, expense_date, expense_type, amount, participants, payer, note
            FROM finance_expenses
            WHERE month = ?
            ORDER BY id
            """,
            (month,)
        ).fetchall()

        incomes = cur.execute(
            """
            SELECT id, income_date, income_type, amount, collector, note
            FROM finance_incomes
            WHERE month = ?
            ORDER BY id
            """,
            (month,)
        ).fetchall()

        members = cur.execute(
            "SELECT id, name, member_type FROM finance_members WHERE month = ? ORDER BY id",
            (month,)
        ).fetchall()

        weekly_map = {
            str(name).strip().lower(): int(weekly or 3)
            for name, weekly in cur.execute("SELECT name, weekly_sessions FROM players").fetchall()
        }

        member_weights = []
        for member_id, name, member_type in members:
            weight = weekly_map.get(str(name).strip().lower(), int(member_type or 3))
            if weight not in [1, 2, 3]:
                weight = 3
            member_weights.append((member_id, name, weight))

        total_weight = sum(int(m[2]) for m in member_weights) or 1
        member_names = [m[1] for m in member_weights]
        member_name_set = {normalize_text_key(n) for n in member_names}
        
        rounded_fund_collector = self.get_auto_income_collector(AUTO_ROUNDED_MEMBER_INCOME, month)
        fund_collector = rounded_fund_collector
        fund_actor_keys = self.get_fund_actor_keys()
        fund_collector_key = normalize_text_key(fund_collector)
        if fund_collector_key:
            fund_actor_keys.add(fund_collector_key)

        # Paid by members
        paid_by_member_map = {name: 0 for name in member_names}
        for _eid, _date, _etype, amount, _participants, payer, _note in expenses:
            payer_key = normalize_text_key(payer)
            if payer_key and payer_key in fund_actor_keys:
                continue
            for member_name in member_names:
                if payer_key == normalize_text_key(member_name):
                    paid_by_member_map[member_name] += int(amount or 0)
                    break

        # Collected by members
        collected_by_member_map = {name: 0 for name in member_names}
        for income_row in incomes:
            collector = str(income_row[4] if len(income_row) > 4 else "").strip()
            amount = int(round(income_row[3] or 0))
            collector_key = normalize_text_key(collector)
            if collector_key and collector_key in fund_actor_keys:
                continue
            for member_name in member_names:
                if collector_key == normalize_text_key(member_name):
                    collected_by_member_map[member_name] += amount
                    break

        # External payments/collections
        paid_by_external_map = {}
        for _eid, _date, _etype, amount, _participants, payer, _note in expenses:
            payer_name = str(payer or "").strip()
            payer_key = normalize_text_key(payer_name)
            if payer_name and payer_key not in fund_actor_keys and payer_key not in member_name_set:
                paid_by_external_map[payer_name] = paid_by_external_map.get(payer_name, 0) + int(amount or 0)

        collected_by_external_map = {}
        for income_row in incomes:
            collector = str((income_row[4] if len(income_row) > 4 else "") or "").strip()
            collector_key = normalize_text_key(collector)
            if collector and collector_key not in fund_actor_keys and collector_key not in member_name_set:
                collected_by_external_map[collector] = collected_by_external_map.get(collector, 0) + int(income_row[3] or 0)

        court_total = sum(int(x[3] or 0) for x in expenses if x[2] == "Thuê sân")
        ball_total = sum(int(x[3] or 0) for x in expenses if x[2] == "Mua bóng")
        picker_total = sum(int(x[3] or 0) for x in expenses if x[2] == "Lượm banh")
        meal_total = sum(int(x[3] or 0) for x in expenses if x[2] == "Ăn uống")
        party_total = sum(int(x[3] or 0) for x in expenses if x[2] in FUND_ONLY_EXPENSE_TYPES)
        other_total = sum(int(x[3] or 0) for x in expenses if x[2] not in KNOWN_EXPENSE_TYPES)
        total_expense = court_total + ball_total + picker_total + meal_total + party_total + other_total
        other_income_total = sum(int(x[3] or 0) for x in incomes)

        extra_share_map = {}
        meal_share_map = {}

        def add_equal_share(target_map, amount, participants):
            if "All" in participants:
                extras = [p for p in participants if p and p != "All"]
                participants = member_names + extras
            participants = [p for p in participants if p and p != "All"]
            if not participants:
                return
            share = int(amount or 0) / len(participants)
            for name in participants:
                target_map[name] = target_map.get(name, 0) + share

        for _, _, expense_type, amount, participants_json, payer, _note in expenses:
            participants = load_json_list(participants_json)

            if expense_type in WEIGHTED_EXPENSE_TYPES:
                if "All" not in participants and participants:
                    add_equal_share(extra_share_map, amount, participants)
                continue

            if expense_type in FUND_ONLY_EXPENSE_TYPES:
                continue

            if expense_type == "Ăn uống":
                add_equal_share(meal_share_map, amount, participants)
            else:
                add_equal_share(extra_share_map, amount, participants)

        # Match fee detail
        all_match_detail_map = self.get_month_match_money_detail_map(month)
        player_name_set = {normalize_text_key(n) for n, _weekly in players}
        match_detail_map = {
            name: data
            for name, data in all_match_detail_map.items()
            if normalize_text_key(name) in player_name_set
        }
        match_money_map = {
            name: data.get("match_money", 0)
            for name, data in match_detail_map.items()
        }
        bet_money_map = {
            name: data.get("bet_money", 0)
            for name, data in match_detail_map.items()
        }
        match_revenue = sum(data.get("match_money", 0) for data in match_detail_map.values())
        bet_revenue = sum(data.get("bet_money", 0) for data in match_detail_map.values())
        prev_match_revenue = match_revenue + bet_revenue

        rows = []

        for member_id, name, weekly in member_weights:
            weight = int(weekly)

            court_share = sum(
                int(x[3] or 0) * weight / total_weight
                for x in expenses
                if x[2] == "Thuê sân" and includes_all_or_empty(x[4])
            )
            ball_picker_share = sum(
                int(x[3] or 0) * weight / total_weight
                for x in expenses
                if x[2] in BALL_PICKER_EXPENSE_TYPES and includes_all_or_empty(x[4])
            )

            meal_share = meal_share_map.get(name, 0)
            other_share = extra_share_map.get(name, 0)

            rows.append({
                "id": member_id,
                "name": name,
                "weekly": weekly,
                "court": court_share,
                "ball_picker": ball_picker_share,
                "meal": meal_share,
                "other_charge": other_share,
            })

        rows.sort(key=lambda r: str(r.get("name","")).lower())

        for r in rows:
            r["match_money"] = match_money_map.get(r["name"], 0)
            r["bet_money"] = bet_money_map.get(r["name"], 0)
            r["calc_fee"] = r["court"] + r["ball_picker"] + r["meal"] + r["other_charge"] + r["match_money"] + r["bet_money"]
            r["rounded_fee"] = int(((r["calc_fee"] + 9999) // 10000) * 10000)

            r["paid_by_member"] = paid_by_member_map.get(r["name"], 0)
            r["collected_by_member"] = collected_by_member_map.get(r["name"], 0)
            r["net_payable"] = r["rounded_fee"] - r["paid_by_member"] + r["collected_by_member"]

        external_names = set()
        for source_map in [meal_share_map, extra_share_map, paid_by_external_map, collected_by_external_map]:
            for name in source_map.keys():
                if str(name).strip() and str(name).strip().lower() not in member_name_set:
                    external_names.add(str(name).strip())

        external_rows = []
        for name in sorted(external_names, key=lambda x: x.lower()):
            meal_value = meal_share_map.get(name, 0)
            other_value = extra_share_map.get(name, 0)
            paid_value = paid_by_external_map.get(name, 0)
            collected_value = collected_by_external_map.get(name, 0)
            calc_value = meal_value + other_value
            net_value = calc_value - paid_value + collected_value
            external_rows.append({
                "name": name,
                "meal": meal_value,
                "other": other_value,
                "paid_by_external": paid_value,
                "collected_by_external": collected_value,
                "net_payable": net_value,
            })

        member_real_collect = sum(r["rounded_fee"] for r in rows)
        external_real_collect = sum(r["net_payable"] for r in external_rows)
        real_collect = member_real_collect + external_real_collect
        need_collect = sum(r["calc_fee"] for r in rows) + sum(r["meal"] + r["other"] for r in external_rows)
        round_extra = sum(r["rounded_fee"] - r["calc_fee"] for r in rows)
        total_rounded_fee = int(round(member_real_collect))
        external_income_total = int(round(external_real_collect))
        total_income_fund = other_income_total + total_rounded_fee + external_income_total
        fund_this_month = total_income_fund - total_expense

        external_fund_collector = self.get_auto_income_collector(AUTO_EXTERNAL_INCOME, month)

        external_fund_collector_key = normalize_text_key(external_fund_collector)
        if (
            external_fund_collector in collected_by_member_map
            and external_fund_collector_key not in fund_actor_keys
        ):
            collected_by_member_map[external_fund_collector] += external_income_total

        if (
            external_fund_collector
            and external_fund_collector_key not in fund_actor_keys
            and external_fund_collector_key not in member_name_set
        ):
            found_external = False
            for er in external_rows:
                if normalize_text_key(er.get("name", "")) == external_fund_collector_key:
                    er["collected_by_external"] = er.get("collected_by_external", 0) + external_income_total
                    er["net_payable"] = er.get("meal", 0) + er.get("other", 0)
                    found_external = True
                    break
            if not found_external and external_income_total:
                external_rows.append({
                    "name": external_fund_collector,
                    "meal": 0,
                    "other": 0,
                    "paid_by_external": 0,
                    "collected_by_external": external_income_total,
                    "net_payable": 0,
                })
            external_real_collect = sum(r.get("net_payable", 0) for r in external_rows)

        if (
            rounded_fund_collector in collected_by_member_map
            and normalize_text_key(rounded_fund_collector) not in fund_actor_keys
        ):
            collected_by_member_map[rounded_fund_collector] += total_rounded_fee

        for r in rows:
            r["collected_by_member"] = collected_by_member_map.get(r["name"], 0)
            r["net_payable"] = (
                r.get("rounded_fee", 0)
                - r.get("paid_by_member", 0)
                + r.get("collected_by_member", 0)
            )

        # Update fund balance in monthly_finance table
        prev_fund = self.get_previous_month_fund(month)
        fund_balance = prev_fund + fund_this_month
        cur.execute(
            "UPDATE monthly_finance SET court_cost=?, ball_cost=?, picker_cost=?, fund_balance=? WHERE month=?",
            (court_total, ball_total, picker_total, fund_balance, month)
        )
        self.conn.commit()

        summary = {
            "month": month,
            "prev_month": previous_month(month),
            "court_total": court_total,
            "ball_total": ball_total,
            "picker_total": picker_total,
            "meal_total": meal_total,
            "party_total": party_total,
            "other_total": other_total,
            "total_expense": total_expense,
            "prev_match_revenue": prev_match_revenue,
            "match_revenue": match_revenue,
            "bet_revenue": bet_revenue,
            "fund_collector": fund_collector,
            "rounded_fund_collector": rounded_fund_collector,
            "other_income_total": other_income_total,
            "total_rounded_fee": total_rounded_fee,
            "external_income_total": external_income_total,
            "total_income_fund": total_income_fund,
            "prev_fund": prev_fund,
            "need_collect": need_collect,
            "real_collect": real_collect,
            "member_real_collect": member_real_collect,
            "external_real_collect": external_real_collect,
            "external_rows": external_rows,
            "round_extra": round_extra,
            "fund_this_month": fund_this_month,
            "fund_balance": fund_balance,
        }
        return rows, expenses, incomes, summary

    def get_month_match_money_detail_map(self, month):
        cur = self.conn.cursor()
        rows = cur.execute(
            """
            SELECT id, match_date, match_no, team_a1, team_a2, team_b1, team_b2,
                   score_a, score_b, match_money, match_bet
            FROM matches
            WHERE substr(match_date, 1, 7) = ?
            ORDER BY match_date, match_no, id
            """,
            (month,)
        ).fetchall()

        rules = self.get_rules()
        money_map = {}

        def get_row(name):
            name = (name or "").strip()
            if not name:
                return None
            if name not in money_map:
                money_map[name] = {"match_money": 0, "bet_money": 0, "total": 0}
            return money_map[name]

        def add_money(name, match_part, bet_part):
            row = get_row(name)
            if row is None:
                return
            row["match_money"] += int(match_part or 0)
            row["bet_money"] += int(bet_part or 0)
            row["total"] += int(match_part or 0) + int(bet_part or 0)

        for _mid, _date, _no, a1, a2, b1, b2, sa, sb, match_money, match_bet in rows:
            team_a = [x for x in [a1, a2] if x]
            team_b = [x for x in [b1, b2] if x]

            base = match_money if match_money is not None else rules["fine_lose"]
            draw = match_money if match_money is not None else rules["fine_draw"]
            zero = match_money if match_money is not None else rules["fine_lose_zero"]
            bet = int(match_bet or 0)

            if sa > sb:
                for name in team_b:
                    add_money(name, zero if sb == 0 else base, bet)
            elif sb > sa:
                for name in team_a:
                    add_money(name, zero if sa == 0 else base, bet)
            else:
                for name in team_a + team_b:
                    add_money(name, draw, bet)

        return money_map


# --- QUẢN LÝ FINANCES CHO CẢ CHU KỲ (THÁNG, QUÝ, NĂM) ---
def calculate_finance_period(db, months):
    aggregated_members = {}
    aggregated_expenses = []
    aggregated_incomes = []
    
    aggregated_summary = {
        "court_total": 0, "ball_total": 0, "picker_total": 0, "meal_total": 0,
        "party_total": 0, "other_total": 0, "total_expense": 0,
        "prev_match_revenue": 0, "match_revenue": 0, "bet_revenue": 0,
        "other_income_total": 0, "total_rounded_fee": 0, "external_income_total": 0,
        "total_income_fund": 0, "prev_fund": 0, "need_collect": 0, "real_collect": 0,
        "member_real_collect": 0, "external_real_collect": 0, "round_extra": 0,
        "fund_this_month": 0, "fund_balance": 0, "external_rows": []
    }
    
    aggregated_external = {}
    
    for m in months:
        rows, expenses, incomes, summary = db.calculate_finance_rows(m)
        aggregated_expenses.extend(expenses)
        aggregated_incomes.extend(incomes)
        
        for r in rows:
            name = r["name"]
            if name not in aggregated_members:
                aggregated_members[name] = {
                    "name": name, "weekly": r["weekly"], "court": 0, "ball_picker": 0,
                    "meal": 0, "other_charge": 0, "match_money": 0, "bet_money": 0,
                    "calc_fee": 0, "rounded_fee": 0, "paid_by_member": 0, "collected_by_member": 0, "net_payable": 0
                }
            m_agg = aggregated_members[name]
            m_agg["court"] += r["court"]
            m_agg["ball_picker"] += r["ball_picker"]
            m_agg["meal"] += r["meal"]
            m_agg["other_charge"] += r["other_charge"]
            m_agg["match_money"] += r["match_money"]
            m_agg["bet_money"] += r["bet_money"]
            m_agg["calc_fee"] += r["calc_fee"]
            m_agg["rounded_fee"] += r["rounded_fee"]
            m_agg["paid_by_member"] += r["paid_by_member"]
            m_agg["collected_by_member"] += r["collected_by_member"]
            m_agg["net_payable"] += r["net_payable"]
            
        for er in summary["external_rows"]:
            name = er["name"]
            if name not in aggregated_external:
                aggregated_external[name] = {
                    "name": name, "meal": 0, "other": 0, "paid_by_external": 0, "collected_by_external": 0, "net_payable": 0
                }
            e_agg = aggregated_external[name]
            e_agg["meal"] += er["meal"]
            e_agg["other"] += er["other"]
            e_agg["paid_by_external"] += er["paid_by_external"]
            e_agg["collected_by_external"] += er["collected_by_external"]
            e_agg["net_payable"] += er["net_payable"]
            
        for k in ["court_total", "ball_total", "picker_total", "meal_total", "party_total", "other_total",
                  "total_expense", "prev_match_revenue", "match_revenue", "bet_revenue",
                  "other_income_total", "total_rounded_fee", "external_income_total",
                  "total_income_fund", "need_collect", "real_collect", "member_real_collect",
                  "external_real_collect", "round_extra", "fund_this_month"]:
            aggregated_summary[k] += summary[k]
            
    sorted_months = sorted(months)
    if sorted_months:
        _, _, _, first_summary = db.calculate_finance_rows(sorted_months[0])
        _, _, _, last_summary = db.calculate_finance_rows(sorted_months[-1])
        aggregated_summary["prev_fund"] = first_summary["prev_fund"]
        aggregated_summary["fund_balance"] = last_summary["fund_balance"]
        
    aggregated_summary["external_rows"] = list(aggregated_external.values())
    return list(aggregated_members.values()), aggregated_expenses, aggregated_incomes, aggregated_summary


# --- TÍNH TOÁN HIỂN THỊ KẾT QUẢ TRẬN ĐẤU DỰA TRÊN LUẬT ---
def get_match_money_value(score_a, score_b, match_money, rules):
    if match_money is not None and not pd.isna(match_money):
        return int(match_money)
    if score_a == score_b:
        return rules["fine_draw"]
    if (score_a > score_b and score_b == 0) or (score_b > score_a and score_a == 0):
        return rules["fine_lose_zero"]
    return rules["fine_lose"]

def get_match_collect_total(team_a1, team_a2, team_b1, team_b2, score_a, score_b, match_money, match_bet, rules):
    money_each = get_match_money_value(score_a, score_b, match_money, rules)
    team_a = [x for x in [team_a1, team_a2] if str(x or "").strip()]
    team_b = [x for x in [team_b1, team_b2] if str(x or "").strip()]
    try:
        sa = int(score_a)
        sb = int(score_b)
    except Exception:
        return 0
    if sa == sb:
        payers = len(team_a) + len(team_b)
    elif sa > sb:
        payers = len(team_b)
    else:
        payers = len(team_a)
    try:
        bet_value = int(match_bet or 0)
    except Exception:
        bet_value = 0
    return (money_each + bet_value) * payers

def get_result_text(score_a, score_b):
    try:
        sa = int(score_a)
        sb = int(score_b)
    except Exception:
        return ""
    if sa > sb:
        return "Đội A thắng"
    if sb > sa:
        return "Đội B thắng"
    return "Hòa"


# --- EXCEL & PDF EXPORTERS ---
def export_excel_bytes(db, months):
    matches = db.fetch_matches(months)
    df_matches = pd.DataFrame(matches, columns=[
        "ID", "Ngày", "Trận số", "A1", "A2", "B1", "B2", "Điểm A", "Điểm B", "Tiền trận", "Tiền độ", "Ghi chú"
    ])
    
    stats = db.calculate_stats(months)
    df_stats = pd.DataFrame(stats)
    if not df_stats.empty:
        df_stats["total_points"] = df_stats["wins"] * db.get_rules()["point_win"] + df_stats["draws"] * db.get_rules()["point_draw"]
        df_stats = df_stats.sort_values(by=["total_points", "wins", "gf"], ascending=[False, False, False])
        df_stats.insert(0, "Hạng", range(1, len(df_stats) + 1))
        df_stats = df_stats.drop(columns=["gf", "ga"], errors="ignore")
        df_stats = df_stats.rename(columns={
            "name": "Tên hội viên", "matches": "Trận", "wins": "Thắng", "draws": "Hòa", "losses": "Thua",
            "match_money": "Tiền trận", "bet_money": "Tiền độ",
            "money": "Tổng tiền", "total_points": "Điểm thành tích"
        })
    
    rows, expenses, incomes, summary = calculate_finance_period(db, months)
    df_members = pd.DataFrame(rows)
    if not df_members.empty:
        df_members = df_members.rename(columns={
            "name": "Hội viên", "weekly": "Buổi/Tuần", "court": "Tiền sân", "ball_picker": "Bóng + Nhặt",
            "meal": "Ăn uống", "other_charge": "Chi khác", "match_money": "Tiền phạt", "bet_money": "Tiền độ",
            "calc_fee": "Phí thực tính", "rounded_fee": "Phí làm tròn", "paid_by_member": "Đã chi hộ",
            "collected_by_member": "Đã thu hộ", "net_payable": "Thu tháng (Cần đóng)"
        })
        
    df_ext = pd.DataFrame(summary["external_rows"])
    if not df_ext.empty:
        df_ext = df_ext.rename(columns={
            "name": "Họ và tên", "meal": "Tiền ăn uống", "other": "Khoản chi khác",
            "paid_by_external": "Đã chi hộ", "collected_by_external": "Đã thu hộ", "net_payable": "Cần thu/trả lại"
        })
        
    df_expenses = pd.DataFrame(expenses, columns=["ID", "Ngày chi", "Loại chi phí", "Số tiền", "Người tham gia", "Người trả", "Ghi chú"])
    if not df_expenses.empty:
        df_expenses["Người tham gia"] = df_expenses["Người tham gia"].apply(lambda x: ", ".join(load_json_list(x)) if "All" not in load_json_list(x) else "Tất cả (All)")
        
    df_incomes = pd.DataFrame(incomes, columns=["ID", "Ngày thu", "Loại khoản thu", "Số tiền", "Người thu", "Ghi chú"])
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_matches.to_excel(writer, sheet_name="Trận đấu", index=False)
        if not df_stats.empty:
            df_stats.to_excel(writer, sheet_name="Xếp hạng", index=False)
        if not df_members.empty:
            df_members.to_excel(writer, sheet_name="Quyết toán Hội viên", index=False)
        if not df_ext.empty:
            df_ext.to_excel(writer, sheet_name="Quyết toán Vãng lai", index=False)
        if not df_expenses.empty:
            df_expenses.to_excel(writer, sheet_name="Khoản Chi", index=False)
        if not df_incomes.empty:
            df_incomes.to_excel(writer, sheet_name="Khoản Thu", index=False)
        
    return output.getvalue()


def export_pdf_bytes(db, months):
    output = BytesIO()
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
        font_name = "HYSMyeongJo-Medium"
    except Exception:
        font_name = "Helvetica"
        
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    
    doc = SimpleDocTemplate(output)
    styles = getSampleStyleSheet()
    elements = []
    
    title_text = f"Báo cáo Tennis Vui - Chu kỳ: {', '.join(months)}"
    title = Paragraph(f"<font name='{font_name}'><b>{title_text}</b></font>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    def add_pdf_table(title_text, df):
        elements.append(Paragraph(f"<font name='{font_name}'><b>{title_text}</b></font>", styles["Heading2"]))
        elements.append(Spacer(1, 6))
        if df.empty:
            elements.append(Paragraph(f"<font name='{font_name}'>Không có dữ liệu</font>", styles["Normal"]))
            elements.append(Spacer(1, 12))
            return
            
        headers = list(df.columns)
        data = [headers]
        for _, row in df.iterrows():
            data.append([str(v) for v in row.values])
            
        tbl = Table(data)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE")
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 16))
        
    matches = db.fetch_matches(months)
    df_matches = pd.DataFrame(matches, columns=[
        "ID", "Ngày", "Trận số", "A1", "A2", "B1", "B2", "Điểm A", "Điểm B", "Tiền trận", "Độ", "Ghi chú"
    ])
    
    stats = db.calculate_stats(months)
    df_stats = pd.DataFrame(stats)
    if not df_stats.empty:
        df_stats["total_points"] = df_stats["wins"] * db.get_rules()["point_win"] + df_stats["draws"] * db.get_rules()["point_draw"]
        df_stats = df_stats.sort_values(by=["total_points", "wins", "gf"], ascending=[False, False, False])
        df_stats.insert(0, "Hạng", range(1, len(df_stats) + 1))
        df_stats = df_stats.drop(columns=["gf", "ga"], errors="ignore")
        df_stats = df_stats.rename(columns={
            "name": "Hội viên", "matches": "Trận", "wins": "T", "draws": "H", "losses": "B",
            "match_money": "Phạt", "bet_money": "Độ", "money": "Tổng", "total_points": "Điểm"
        })
        
    rows, expenses, incomes, summary = calculate_finance_period(db, months)
    df_members = pd.DataFrame(rows)
    if not df_members.empty:
        df_members = df_members.rename(columns={
            "name": "Hội viên", "weekly": "Buổi", "court": "Sân", "ball_picker": "Bóng",
            "meal": "Ăn", "other_charge": "Khác", "match_money": "Phạt", "bet_money": "Độ",
            "calc_fee": "Thực", "rounded_fee": "Tròn", "paid_by_member": "Chi hộ",
            "collected_by_member": "Thu hộ", "net_payable": "Thu tháng"
        })
        df_members = df_members[["Hội viên", "Buổi", "Sân", "Bóng", "Ăn", "Khác", "Phạt", "Độ", "Chi hộ", "Thu hộ", "Thu tháng"]]
        
    df_expenses = pd.DataFrame(expenses, columns=["ID", "Ngày", "Loại", "Số tiền", "Tham gia", "Người chi", "Ghi chú"])
    if not df_expenses.empty:
        df_expenses["Tham gia"] = df_expenses["Tham gia"].apply(lambda x: ", ".join(load_json_list(x)) if "All" not in load_json_list(x) else "All")
        df_expenses = df_expenses[["ID", "Ngày", "Loại", "Số tiền", "Tham gia", "Người chi"]]

    add_pdf_table("Thống kê trận đấu", df_matches)
    if not df_stats.empty:
        add_pdf_table("Bảng xếp hạng", df_stats)
    if not df_members.empty:
        add_pdf_table("Quyết toán hội viên", df_members)
    if not df_expenses.empty:
        add_pdf_table("Danh sách khoản chi", df_expenses)
    
    doc.build(elements)
    return output.getvalue()


# --- EXCEL IMPORTER ---
def import_excel_data(uploaded_file, db):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return False, "File Excel không có dữ liệu."
            
        header_aliases = {
            "match_date": {"ngay", "ngay_danh", "ngay_thi_dau", "date", "match_date"},
            "match_no": {"tran", "tran_so", "so_tran", "match_no"},
            "team_a1": {"a1", "doi_a1", "team_a1"},
            "team_a2": {"a2", "doi_a2", "team_a2"},
            "score_a": {"a", "ty_so_a", "diem_a", "score_a"},
            "score_b": {"b", "ty_so_b", "diem_b", "score_b"},
            "team_b1": {"b1", "doi_b1", "team_b1"},
            "team_b2": {"b2", "doi_b2", "team_b2"},
            "match_money": {"tien_tran", "tien", "match_money", "money"},
            "match_bet": {"tien_do", "do", "match_bet", "bet"},
            "note": {"ghi_chu", "ghichu", "note", "remark"},
        }
        
        default_order = [
            "match_date", "match_no", "team_a1", "team_a2",
            "score_a", "score_b", "team_b1", "team_b2", "match_money"
        ]
        
        first_row = rows[0]
        normalized = [normalize_text_key(v) for v in first_row]
        col_map = {}
        
        for idx, h in enumerate(normalized):
            for field, aliases in header_aliases.items():
                if h in aliases:
                    col_map[field] = idx
                    
        has_header = all(k in col_map for k in ["match_date", "team_a1", "team_b1", "score_a", "score_b"])
        
        data_rows = rows[1:] if has_header else rows
        if not has_header:
            col_map = {field: idx for idx, field in enumerate(default_order)}
            
        def cell(row, field, default=""):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return "" if value is None else value
            
        def parse_date_value(value):
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            text_value = str(value).strip()
            if not text_value:
                raise ValueError("Thiếu ngày")
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(text_value, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
            raise ValueError(f"Ngày '{text_value}' phải có dạng YYYY-MM-DD hoặc DD/MM/YYYY")
            
        def parse_money(val):
            if val is None or str(val).strip() == "":
                return None
            try:
                txt = str(val).replace(",", "").replace(".", "").replace("đ", "").replace(" ", "")
                return int(txt)
            except Exception:
                return None

        count = 0
        errors = []
        cur = db.conn.cursor()
        
        for excel_row_no, row in enumerate(data_rows, start=2 if has_header else 1):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
                
            try:
                match_date = parse_date_value(cell(row, "match_date"))
                
                match_no_raw = cell(row, "match_no", "")
                if str(match_no_raw).strip() == "":
                    match_no = cur.execute(
                        "SELECT COALESCE(MAX(match_no), 0) + 1 FROM matches WHERE match_date = ?",
                        (match_date,)
                    ).fetchone()[0]
                else:
                    match_no = int(float(str(match_no_raw).strip()))
                    
                money_raw = cell(row, "match_money", "")
                match_money = parse_money(money_raw)
                
                bet_raw = cell(row, "match_bet", "")
                match_bet = parse_money(bet_raw) if bet_raw else 0
                
                data = {
                    "match_date": match_date,
                    "match_no": match_no,
                    "team_a1": str(cell(row, "team_a1")).strip(),
                    "team_a2": str(cell(row, "team_a2")).strip(),
                    "team_b1": str(cell(row, "team_b1")).strip(),
                    "team_b2": str(cell(row, "team_b2")).strip(),
                    "score_a": int(float(str(cell(row, "score_a")).strip())),
                    "score_b": int(float(str(cell(row, "score_b")).strip())),
                    "match_money": match_money,
                    "match_bet": match_bet,
                    "note": str(cell(row, "note", "")).strip(),
                }
                
                db.validate_match_data(data)
                cur.execute(
                    """
                    INSERT INTO matches(match_date, match_no, team_a1, team_a2, team_b1, team_b2, score_a, score_b, match_money, match_bet, note, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["match_date"], data["match_no"], data["team_a1"], data["team_a2"], data["team_b1"], data["team_b2"],
                        data["score_a"], data["score_b"], data["match_money"], data["match_bet"], data["note"], datetime.now().isoformat(timespec="seconds")
                    )
                )
                count += 1
                
            except Exception as e:
                errors.append(f"Dòng {excel_row_no}: {e}")
                
        db.conn.commit()
        if errors:
            return True, f"Đã nhập {count} trận. Có {len(errors)} dòng lỗi:\n" + "\n".join(errors[:5])
        return True, f"Đã nhập thành công {count} trận từ Excel."
    except Exception as e:
        return False, f"Lỗi đọc file: {e}"


# --- INLINE GRID EDITOR SAVE LOGIC ---
def save_editor_changes(db, table_name, editor_key, df_original):
    if editor_key not in st.session_state:
        return True, "Không có thay đổi nào."
    
    changes = st.session_state[editor_key]
    if not changes or (not changes.get("edited_rows") and not changes.get("added_rows") and not changes.get("deleted_rows")):
        return True, "Không có thay đổi nào."
    
    cur = db.conn.cursor()
    try:
        # 1. Deletions
        if "deleted_rows" in changes and changes["deleted_rows"]:
            for row_idx in changes["deleted_rows"]:
                row_id_val = df_original.iloc[row_idx]["ID"]
                if row_id_val == "AUTO" or str(row_id_val).strip() == "AUTO":
                    continue
                row_id = int(row_id_val)
                cur.execute(f"DELETE FROM {table_name} WHERE id = ?", (row_id,))
        
        # 2. Edits
        if "edited_rows" in changes and changes["edited_rows"]:
            for row_idx_str, col_changes in changes["edited_rows"].items():
                row_idx = int(row_idx_str)
                row_id_val = df_original.iloc[row_idx]["ID"]
                if row_id_val == "AUTO" or str(row_id_val).strip() == "AUTO":
                    if table_name == "finance_incomes":
                        new_collector = col_changes.get("Người thu")
                        if new_collector:
                            income_type = df_original.iloc[row_idx]["Loại khoản thu"]
                            month = df_original.iloc[row_idx]["Ngày thu"]
                            key = f"auto_income_collector::{month}::{income_type}"
                            cur.execute("INSERT OR REPLACE INTO settings(key, value) VALUES (?, ?)", (key, new_collector))
                    continue
                row_id = int(row_id_val)
                row_data = df_original.iloc[row_idx].to_dict()
                row_data.update(col_changes)
                
                if table_name == "players":
                    name = str(row_data["Họ và tên"]).strip()
                    if not name:
                        raise ValueError("Tên hội viên không được để trống!")
                    active = 1 if row_data["Hoạt động"] else 0
                    weekly = int(row_data["Số buổi/Tuần"])
                    cur.execute("UPDATE players SET name = ?, active = ?, weekly_sessions = ? WHERE id = ?", (name, active, weekly, row_id))
                
                elif table_name == "matches":
                    money_val = row_data["Tiền trận"]
                    match_money = int(money_val) if money_val is not None and not pd.isna(money_val) else None
                    bet_val = row_data["Tiền độ"]
                    match_bet = int(bet_val) if bet_val is not None and not pd.isna(bet_val) else 0
                    
                    data = {
                        "match_date": str(row_data["Ngày"]).strip(),
                        "match_no": int(row_data["Trận"]),
                        "team_a1": str(row_data["A1"]).strip(),
                        "team_a2": str(row_data["A2"]).strip() if not pd.isna(row_data["A2"]) else "",
                        "team_b1": str(row_data["B1"]).strip(),
                        "team_b2": str(row_data["B2"]).strip() if not pd.isna(row_data["B2"]) else "",
                        "score_a": int(row_data["A"]),
                        "score_b": int(row_data["B"]),
                        "match_money": match_money,
                        "match_bet": match_bet,
                        "note": str(row_data["Ghi chú"]).strip() if not pd.isna(row_data["Ghi chú"]) else "",
                    }
                    db.validate_match_data(data)
                    cur.execute(
                        """
                        UPDATE matches SET match_date = ?, match_no = ?, team_a1 = ?, team_a2 = ?, team_b1 = ?, team_b2 = ?,
                                           score_a = ?, score_b = ?, match_money = ?, match_bet = ?, note = ?
                        WHERE id = ?
                        """,
                        (data["match_date"], data["match_no"], data["team_a1"], data["team_a2"], data["team_b1"], data["team_b2"],
                         data["score_a"], data["score_b"], data["match_money"], data["match_bet"], data["note"], row_id)
                    )
                
                elif table_name == "finance_expenses":
                    amount = int(row_data["Số tiền"])
                    payer = str(row_data["Người trả"]).strip()
                    note = str(row_data["Ghi chú"]).strip() if not pd.isna(row_data["Ghi chú"]) else ""
                    expense_type = str(row_data["Loại chi phí"]).strip()
                    expense_date = str(row_data["Ngày chi"]).strip()
                    
                    parts_text = str(row_data["Người tham gia"]).strip()
                    if parts_text.lower() == "tất cả (all)" or parts_text.lower() == "all":
                        parts = ["All"]
                    else:
                        parts = [p.strip() for p in parts_text.split(",") if p.strip()]
                    parts_json = json.dumps(parts)
                    
                    cur.execute(
                        """
                        UPDATE finance_expenses SET expense_date = ?, expense_type = ?, amount = ?, participants = ?, payer = ?, note = ?
                        WHERE id = ?
                        """,
                        (expense_date, expense_type, amount, parts_json, payer, note, row_id)
                    )
                
                elif table_name == "finance_incomes":
                    amount = int(row_data["Số tiền"])
                    collector = str(row_data["Người thu"]).strip()
                    note = str(row_data["Ghi chú"]).strip() if not pd.isna(row_data["Ghi chú"]) else ""
                    income_type = str(row_data["Loại khoản thu"]).strip()
                    income_date = str(row_data["Ngày thu"]).strip()
                    
                    cur.execute(
                        """
                        UPDATE finance_incomes SET income_date = ?, income_type = ?, amount = ?, collector = ?, note = ?
                        WHERE id = ?
                        """,
                        (income_date, income_type, amount, collector, note, row_id)
                    )
        
        # 3. Additions
        if "added_rows" in changes and changes["added_rows"]:
            for row_data in changes["added_rows"]:
                if table_name == "players":
                    name = str(row_data.get("Họ và tên", "")).strip()
                    if not name:
                        raise ValueError("Tên hội viên không được để trống!")
                    active = 1 if row_data.get("Hoạt động", True) else 0
                    weekly = int(row_data.get("Số buổi/Tuần", 3))
                    cur.execute("INSERT INTO players(name, active, weekly_sessions) VALUES (?, ?, ?)", (name, active, weekly))
                
                elif table_name == "matches":
                    money_val = row_data.get("Tiền trận")
                    match_money = int(money_val) if money_val is not None and not pd.isna(money_val) else None
                    bet_val = row_data.get("Tiền độ")
                    match_bet = int(bet_val) if bet_val is not None and not pd.isna(bet_val) else 0
                    
                    match_date_val = str(row_data.get("Ngày", datetime.now().strftime("%Y-%m-%d"))).strip()
                    match_no_val = row_data.get("Trận")
                    if match_no_val is None or pd.isna(match_no_val):
                        max_no = cur.execute("SELECT COALESCE(MAX(match_no), 0) FROM matches WHERE match_date = ?", (match_date_val,)).fetchone()[0]
                        match_no = max_no + 1
                    else:
                        match_no = int(match_no_val)
                    
                    data = {
                        "match_date": match_date_val,
                        "match_no": match_no,
                        "team_a1": str(row_data.get("A1", "")).strip(),
                        "team_a2": str(row_data.get("A2", "")).strip(),
                        "team_b1": str(row_data.get("B1", "")).strip(),
                        "team_b2": str(row_data.get("B2", "")).strip(),
                        "score_a": int(row_data.get("A", 0)),
                        "score_b": int(row_data.get("B", 0)),
                        "match_money": match_money,
                        "match_bet": match_bet,
                        "note": str(row_data.get("Ghi chú", "")).strip(),
                    }
                    db.validate_match_data(data)
                    cur.execute(
                        """
                        INSERT INTO matches (match_date, match_no, team_a1, team_a2, team_b1, team_b2, score_a, score_b, match_money, match_bet, note, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (data["match_date"], data["match_no"], data["team_a1"], data["team_a2"], data["team_b1"], data["team_b2"],
                         data["score_a"], data["score_b"], data["match_money"], data["match_bet"], data["note"], datetime.now().isoformat(timespec="seconds"))
                    )
                
                elif table_name == "finance_expenses":
                    amount = int(row_data.get("Số tiền", 0))
                    payer = str(row_data.get("Người trả", "Quỹ")).strip()
                    note = str(row_data.get("Ghi chú", "")).strip()
                    expense_type = str(row_data.get("Loại chi phí", EXPENSE_TYPES[0])).strip()
                    expense_date = str(row_data.get("Ngày chi", datetime.now().strftime("%Y-%m-%d"))).strip()
                    
                    parts_text = str(row_data.get("Người tham gia", "All")).strip()
                    if parts_text.lower() == "all" or not parts_text:
                        parts = ["All"]
                    else:
                        parts = [p.strip() for p in parts_text.split(",") if p.strip()]
                    parts_json = json.dumps(parts)
                    month = expense_date[:7]
                    
                    cur.execute(
                        """
                        INSERT INTO finance_expenses (month, expense_date, expense_type, amount, participants, payer, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (month, expense_date, expense_type, amount, parts_json, payer, note)
                    )
                
                elif table_name == "finance_incomes":
                    amount = int(row_data.get("Số tiền", 0))
                    collector = str(row_data.get("Người thu", "Quỹ")).strip()
                    note = str(row_data.get("Ghi chú", "")).strip()
                    income_type = str(row_data.get("Loại khoản thu", DEFAULT_INCOME_TYPES[0])).strip()
                    income_date = str(row_data.get("Ngày thu", datetime.now().strftime("%Y-%m-%d"))).strip()
                    month = income_date[:7]
                    
                    cur.execute(
                        """
                        INSERT INTO finance_incomes (month, income_date, income_type, amount, collector, note)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (month, income_date, income_type, amount, collector, note)
                    )
        
        db.conn.commit()
        return True, "Đã lưu tất cả thay đổi thành công!"
    except Exception as e:
        db.conn.rollback()
        return False, f"Lỗi: {e}"


# --- GIAO DIỆN STREAMLIT ---
# Tùy chỉnh CSS giao diện
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&family=Inter:wght@300;400;500;600;700&display=swap');
    
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Outfit', 'Inter', sans-serif;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #1e293b;
        border-radius: 6px;
        color: #94a3b8;
        font-weight: 600;
        padding: 8px 16px;
        border: 1px solid #334155;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background-color: #334155;
        color: #f8fafc;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #0ea5e9 0%, #0284c7 100%) !important;
        color: #ffffff !important;
        border: none !important;
    }
    
    .metric-card {
        background: #1e293b;
        border: 1px solid #334155;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1);
        text-align: center;
    }
    
    .metric-title {
        color: #94a3b8;
        font-size: 14px;
        font-weight: 500;
        margin-bottom: 5px;
    }
    
    .metric-value {
        font-size: 24px;
        font-weight: 700;
        margin-bottom: 5px;
    }
    
    .header-gradient {
        background: linear-gradient(90deg, #38bdf8 0%, #0284c7 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-weight: 800;
        font-size: 42px;
        margin-bottom: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# --- KIỂM TRA ĐĂNG NHẬP ---
def get_cookies():
    try:
        return st.context.cookies
    except Exception:
        return {}

# Xóa cookie nếu có yêu cầu đăng xuất
if st.session_state.get("delete_login_cookie"):
    st.html(
        """
        <script>
            document.cookie = "tennis_logged_in=; max-age=0; path=/; SameSite=Lax";
        </script>
        """,
        unsafe_allow_javascript=True
    )
    del st.session_state["delete_login_cookie"]

# Ghi cookie nếu có yêu cầu lưu đăng nhập
if st.session_state.get("write_login_cookie"):
    st.html(
        """
        <script>
            document.cookie = "tennis_logged_in=true; max-age=2592000; path=/; SameSite=Lax";
        </script>
        """,
        unsafe_allow_javascript=True
    )
    del st.session_state["write_login_cookie"]

# Kiểm tra trạng thái đăng nhập từ cookie
cookies = get_cookies()
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = (cookies.get("tennis_logged_in") == "true")

if not st.session_state["logged_in"]:
    st.markdown('<div class="header-gradient" style="text-align: center; margin-top: 80px;">TENNIS VUI</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            st.subheader("🔑 Đăng nhập hệ thống")
            username = st.text_input("Tên đăng nhập", value="admin", placeholder="Nhập tên đăng nhập...")
            password = st.text_input("Mật khẩu", type="password", placeholder="Nhập mật khẩu...")
            remember_me = st.checkbox("Ghi nhớ đăng nhập trên thiết bị này")
            submitted = st.form_submit_button("Đăng nhập", type="primary", use_container_width=True)
            
            if submitted:
                if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                    st.session_state["logged_in"] = True
                    if remember_me:
                        st.session_state["write_login_cookie"] = True
                    st.success("Đăng nhập thành công!")
                    st.rerun()
                else:
                    st.error("Sai tên đăng nhập hoặc mật khẩu!")
    st.stop()

# --- QUẢN LÝ DANH SÁCH FILE DATABASE ---
st.sidebar.markdown("### 🎾 CƠ SỞ DỮ LIỆU")

# Tìm kiếm các file .tennis trong thư mục
local_files = [f for f in os.listdir(".") if f.endswith(".tennis")]
if DEFAULT_DB_FILE not in local_files and not os.path.exists(DEFAULT_DB_FILE):
    conn = sqlite3.connect(DEFAULT_DB_FILE)
    conn.close()
    local_files.append(DEFAULT_DB_FILE)

# Cho phép tải file lên
uploaded_file = st.sidebar.file_uploader("Tải file .tennis từ máy lên", type=["tennis", "db"])
if uploaded_file is not None:
    save_path = Path(uploaded_file.name)
    with open(save_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    st.sidebar.success(f"Đã tải lên: {uploaded_file.name}")
    if uploaded_file.name not in local_files:
        local_files.append(uploaded_file.name)

# Chọn database hoạt động
selected_db = st.sidebar.selectbox(
    "Chọn file dữ liệu để làm việc:",
    options=sorted(local_files),
    index=sorted(local_files).index(DEFAULT_DB_FILE) if DEFAULT_DB_FILE in local_files else 0
)

# Tự động tạo bản sao lưu hàng ngày (giới hạn tối đa 5 bản)
perform_auto_daily_backup(selected_db)

# Khởi tạo db kết nối (mỗi lần rerun sẽ mở kết nối riêng để tránh lỗi đa luồng)
db = TennisDB(selected_db)

# --- TRÍCH XUẤT THÁNG, QUÝ, NĂM TỪ DATABASE ---
def get_time_periods(db):
    cur = db.conn.cursor()
    months = set()
    try:
        for r in cur.execute("SELECT DISTINCT substr(match_date, 1, 7) FROM matches").fetchall():
            if r[0] and len(r[0]) == 7: months.add(r[0])
    except Exception: pass
    try:
        for r in cur.execute("SELECT DISTINCT month FROM finance_expenses").fetchall():
            if r[0] and len(r[0]) == 7: months.add(r[0])
    except Exception: pass
    try:
        for r in cur.execute("SELECT DISTINCT month FROM finance_incomes").fetchall():
            if r[0] and len(r[0]) == 7: months.add(r[0])
    except Exception: pass
    try:
        for r in cur.execute("SELECT DISTINCT month FROM monthly_finance").fetchall():
            if r[0] and len(r[0]) == 7: months.add(r[0])
    except Exception: pass

    if not months:
        months.add(datetime.now().strftime("%Y-%m"))
        
    months_list = sorted(list(months), reverse=True)
    
    quarters = set()
    years = set()
    for m in months_list:
        try:
            y, mo = m.split("-")
            years.add(y)
            q = (int(mo) - 1) // 3 + 1
            quarters.add(f"{y}-Q{q}")
        except Exception:
            pass
            
    return months_list, sorted(list(quarters), reverse=True), sorted(list(years), reverse=True)

# Lựa chọn khoảng thời gian lọc dữ liệu
period_type = st.sidebar.selectbox("Lọc dữ liệu theo:", ["Tháng", "Quý", "Năm"])
months_list, quarters_list, years_list = get_time_periods(db)

if period_type == "Tháng":
    selected_period = st.sidebar.selectbox("Chọn tháng:", options=months_list)
    months_to_query = [selected_period]
elif period_type == "Quý":
    selected_period = st.sidebar.selectbox("Chọn quý:", options=quarters_list)
    y, q = selected_period.split("-Q")
    q_num = int(q)
    start_mo = (q_num - 1) * 3 + 1
    months_to_query = [f"{y}-{mo:02d}" for mo in range(start_mo, start_mo + 3)]
else:
    selected_period = st.sidebar.selectbox("Chọn năm:", options=years_list)
    months_to_query = [f"{selected_period}-{mo:02d}" for mo in range(1, 13)]

# Nút sao lưu nhanh database
if st.sidebar.button("💾 Sao lưu (Backup)"):
    backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{selected_db}"
    shutil.copy(selected_db, backup_name)
    st.sidebar.success(f"Đã sao lưu thành: {backup_name}")

# Nút tải trực tiếp file dữ liệu về thiết bị (để lưu trữ vào OneDrive / Google Drive)
try:
    with open(selected_db, "rb") as f:
        db_bytes = f.read()
    st.sidebar.download_button(
        label="📥 Tải file dữ liệu (.tennis)",
        data=db_bytes,
        file_name=selected_db,
        mime="application/octet-stream",
        use_container_width=True
    )
except Exception as e:
    st.sidebar.error(f"Lỗi đọc file: {e}")

# Nút xóa file dữ liệu đang chọn
with st.sidebar.popover("🗑️ Xóa file dữ liệu", use_container_width=True):
    file_to_delete = st.selectbox("Chọn file muốn xóa:", options=sorted(local_files), key="file_to_delete_sb")
    st.warning(f"⚠️ Hành động này sẽ xóa vĩnh viễn file `{file_to_delete}` khỏi hệ thống.")
    confirm_delete = st.checkbox("Tôi chắc chắn muốn xóa file này", key="confirm_delete_db")
    if st.button("Xác nhận Xóa Vĩnh Viễn", type="primary", key="delete_db_btn", disabled=not confirm_delete, use_container_width=True):
        try:
            if file_to_delete == selected_db:
                db.conn.close()
            os.remove(file_to_delete)
            st.success(f"Đã xóa file `{file_to_delete}` thành công!")
            st.rerun()
        except Exception as e:
            st.error(f"Lỗi khi xóa file: {e}")

# Nút đăng xuất ở cuối sidebar
st.sidebar.write("---")

# Nút lưu trữ/đồng bộ lên đám mây GitHub
if "GITHUB_TOKEN" in st.secrets and "GITHUB_REPO" in st.secrets:
    col_git1, col_git2 = st.sidebar.columns(2)
    with col_git1:
        with st.popover("☁️ Lưu lên GitHub", use_container_width=True):
            github_name = st.text_input("Tên file trên GitHub:", value=selected_db, key="gh_save_name")
            if st.button("Xác nhận Lưu", type="primary", use_container_width=True, key="sync_push_btn"):
                if not github_name.strip():
                    st.error("Tên file không được để trống!")
                else:
                    with st.spinner("Đang lưu lên..."):
                        success, msg = save_db_to_github(selected_db, github_name.strip())
                        if success:
                            st.success(msg)
                        else:
                            st.error(msg)
    with col_git2:
        with st.popover("☁️ Tải từ GitHub", use_container_width=True):
            if "github_files" not in st.session_state:
                st.session_state["github_files"] = []
                
            if st.button("🔄 Lấy danh sách file", key="refresh_gh_files_btn", use_container_width=True):
                with st.spinner("Đang lấy danh sách..."):
                    files = list_files_on_github()
                    if files:
                        st.session_state["github_files"] = files
                        st.success(f"Đã tìm thấy {len(files)} file!")
                    else:
                        st.warning("Không tìm thấy file nào hoặc lỗi kết nối!")
            
            if st.session_state["github_files"]:
                github_selected_file = st.selectbox(
                    "Chọn file trên GitHub:",
                    options=st.session_state["github_files"],
                    key="gh_pull_select"
                )
            else:
                github_selected_file = st.text_input(
                    "Nhập tên file trên GitHub:",
                    value=selected_db,
                    key="gh_pull_text"
                )
                
            local_save_name = st.text_input(
                "Lưu thành file local:",
                value=github_selected_file if github_selected_file else selected_db,
                key="gh_pull_local_name"
            )
            
            if st.button("Xác nhận Tải", type="primary", use_container_width=True, key="gh_pull_confirm_btn"):
                if not github_selected_file.strip() or not local_save_name.strip():
                    st.error("Tên file không được để trống!")
                else:
                    with st.spinner("Đang tải..."):
                        success, msg = load_db_from_github(github_selected_file.strip(), local_save_name.strip())
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
else:
    with st.sidebar.expander("☁️ Đồng bộ GitHub (Đám mây)"):
        st.info(
            "Để đồng bộ 2 chiều với GitHub từ web, hãy thêm các cấu hình sau vào phần **Secrets** của Streamlit Cloud:\n"
            "```toml\n"
            "GITHUB_TOKEN = \"your_personal_access_token\"\n"
            "GITHUB_REPO = \"your_username/your_repo_name\"\n"
            "```"
        )

st.sidebar.write("---")
if st.sidebar.button("🚗 Đăng xuất (Logout)", use_container_width=True, key="logout_btn"):
    st.session_state["logged_in"] = False
    st.session_state["delete_login_cookie"] = True
    st.rerun()

# Tiêu đề chính
st.markdown(f'<div class="header-gradient">TENNIS VUI</div>', unsafe_allow_html=True)
st.markdown(f"**Dữ liệu đang xem:** `{selected_db}` | **Khoảng thời gian ({period_type}):** `{selected_period}`")

# Tính toán các dòng tài chính của thời gian hiện tại
member_finance_rows, expenses_list, incomes_list, finance_summary = calculate_finance_period(db, months_to_query)
players_list = db.get_player_names()
rules = db.get_rules()

# --- THIẾT LẬP CÁC TAB CHÍNH (ĐÚNG THỨ TỰ BẢN GỐC) ---
# 1. Nhập trận | 2. Bảng xếp hạng | 3. Bảng tiền trận | 4. Thống kê trận | 5. Thu chi tháng | 6. Dữ liệu
tab_input_match, tab_rank, tab_money, tab_stats, tab_finance, tab_data = st.tabs([
    "📝 Nhập trận", 
    "🏆 Bảng xếp hạng", 
    "💵 Bảng tiền trận", 
    "📊 Thống kê trận", 
    "💰 Thu chi tháng", 
    "⚙️ Dữ liệu"
])

# =========================================================
# TAB 1: NHẬP TRẬN (INPUT MATCH)
# =========================================================
with tab_input_match:
    # Form thêm trận đấu mới nhanh (giống y hệt giao diện gốc)
    st.subheader("Nhập kết quả trận đấu")
    col_f1, col_f2, col_f3 = st.columns(3)
    
    with col_f1:
        match_date = st.date_input("Ngày thi đấu", value=datetime.now(), key="in_date")
        
        # Đội A selectbox
        a1_select = st.selectbox("Đội A - Người 1", options=players_list + ["-- Nhập người ngoài --"], key="in_a1")
        if a1_select == "-- Nhập người ngoài --":
            a1 = st.text_input("Đội A - Nhập tên người ngoài 1", key="in_a1_custom").strip()
        else:
            a1 = a1_select
            
        a2_select = st.selectbox("Đội A - Người 2 (Đơn để trống)", options=["", "-- Nhập người ngoài --"] + players_list, key="in_a2")
        if a2_select == "-- Nhập người ngoài --":
            a2 = st.text_input("Đội A - Nhập tên người ngoài 2", key="in_a2_custom").strip()
        else:
            a2 = a2_select
            
    with col_f2:
        # Tự động tính trận số
        d_str = match_date.strftime("%Y-%m-%d")
        cur = db.conn.cursor()
        max_no = cur.execute("SELECT COALESCE(MAX(match_no), 0) FROM matches WHERE match_date=?", (d_str,)).fetchone()[0]
        match_no = st.number_input("Trận số", min_value=1, value=max_no + 1, key="in_no")
        
        # Đội B selectbox
        b1_select = st.selectbox("Đội B - Người 1", options=players_list + ["-- Nhập người ngoài --"], key="in_b1")
        if b1_select == "-- Nhập người ngoài --":
            b1 = st.text_input("Đội B - Nhập tên người ngoài 1", key="in_b1_custom").strip()
        else:
            b1 = b1_select
            
        b2_select = st.selectbox("Đội B - Người 2 (Đơn để trống)", options=["", "-- Nhập người ngoài --"] + players_list, key="in_b2")
        if b2_select == "-- Nhập người ngoài --":
            b2 = st.text_input("Đội B - Nhập tên người ngoài 2", key="in_b2_custom").strip()
        else:
            b2 = b2_select
            
    with col_f3:
        custom_fine = st.checkbox("Tiền trận", key="in_custom_fine")
        match_money_val = None
        if custom_fine:
            match_money_val = st.number_input("Tiền trận", min_value=0, step=5000, value=20000, key="in_money")
        else:
            st.info("💡 Mặc định sử dụng Luật cấu hình")
            
        match_bet = st.number_input("Tiền độ", min_value=0, step=10000, value=0, key="in_bet")
        
        # Nhập tỷ số
        col_sa, col_sb = st.columns(2)
        with col_sa:
            score_a = st.number_input("Tỷ số Đội A", min_value=0, max_value=7, value=0, key="in_score_a")
        with col_sb:
            score_b = st.number_input("Tỷ số Đội B", min_value=0, max_value=7, value=0, key="in_score_b")
            
    note = st.text_input("Ghi chú trận", key="in_note")
    
    col_btn1, col_btn2, _ = st.columns([1, 1, 6])
    with col_btn1:
        if st.button("Thêm trận", key="in_save_btn", type="primary"):
            try:
                data = {
                    "match_date": d_str,
                    "match_no": int(match_no),
                    "team_a1": a1,
                    "team_a2": a2 or "",
                    "team_b1": b1,
                    "team_b2": b2 or "",
                    "score_a": int(score_a),
                    "score_b": int(score_b),
                    "match_money": match_money_val,
                    "match_bet": int(match_bet),
                    "note": note,
                }
                db.validate_match_data(data)
                db.conn.execute(
                    """
                    INSERT INTO matches (match_date, match_no, team_a1, team_a2, team_b1, team_b2, score_a, score_b, match_money, created_at, match_bet, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (data["match_date"], data["match_no"], data["team_a1"], data["team_a2"], data["team_b1"], data["team_b2"],
                     data["score_a"], data["score_b"], data["match_money"], datetime.now().isoformat(timespec="seconds"), data["match_bet"], data["note"])
                )
                db.conn.commit()
                st.success("Đã thêm trận đấu thành công!")
                st.rerun()
            except Exception as e:
                st.error(f"Lỗi: {e}")
                
    with col_btn2:
        # File excel uploader to import matches
        with st.popover("📥 Nhập Excel"):
            excel_file = st.file_uploader("Chọn tệp Excel (.xlsx)", type=["xlsx"])
            if excel_file is not None:
                if st.button("Bắt đầu Nhập từ Excel"):
                    success, msg = import_excel_data(excel_file, db)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                        
    st.write("---")

    # Lưới chỉnh sửa matches
    st.subheader("Danh sách kết quả trận đấu")
    matches_raw = db.fetch_matches(months_to_query)
    if matches_raw:
        # Chuẩn bị Dataframe phù hợp hiển thị bản gốc
        df_matches = pd.DataFrame(matches_raw, columns=[
            "ID", "Ngày", "Trận", "A1", "A2", "B1", "B2", "A", "B", "Tiền trận", "Tiền độ", "Ghi chú"
        ])
        
        # Sắp xếp lại cột để giống Treeview gốc
        df_matches = df_matches[["ID", "Ngày", "Trận", "A1", "A2", "A", "B", "B1", "B2", "Tiền trận", "Tiền độ", "Ghi chú"]]
        
        # Tính toán các cột động hiển thị
        df_matches["Tiền thu trận"] = df_matches.apply(
            lambda r: get_match_collect_total(r["A1"], r["A2"], r["B1"], r["B2"], r["A"], r["B"], r["Tiền trận"], r["Tiền độ"], rules), axis=1
        )
        df_matches["Kết quả"] = df_matches.apply(
            lambda r: get_result_text(r["A"], r["B"]), axis=1
        )
        
        # Sắp xếp thứ tự cột hiển thị
        cols_order = ["ID", "Ngày", "Trận", "A1", "A2", "A", "B", "B1", "B2", "Tiền trận", "Tiền độ", "Tiền thu trận", "Kết quả", "Ghi chú"]
        df_matches = df_matches[cols_order]
        
        edited_matches_df = st.data_editor(
            df_matches,
            num_rows="dynamic",
            key="matches_editor",
            column_config={
                "ID": st.column_config.NumberColumn("ID", disabled=True, format="%d"),
                "Ngày": st.column_config.TextColumn("Ngày"),
                "Trận": st.column_config.NumberColumn("Trận", format="%d"),
                "A1": st.column_config.TextColumn("A1"),
                "A2": st.column_config.TextColumn("A2"),
                "A": st.column_config.NumberColumn("A", min_value=0, max_value=7, format="%d"),
                "B": st.column_config.NumberColumn("B", min_value=0, max_value=7, format="%d"),
                "B1": st.column_config.TextColumn("B1"),
                "B2": st.column_config.TextColumn("B2"),
                "Tiền trận": st.column_config.NumberColumn("Tiền trận", format="%,.0f"),
                "Tiền độ": st.column_config.NumberColumn("Tiền độ", format="%,.0f"),
                "Tiền thu trận": st.column_config.NumberColumn("Tiền thu trận", disabled=True, format="%,.0f"),
                "Kết quả": st.column_config.TextColumn("Kết quả", disabled=True),
                "Ghi chú": st.column_config.TextColumn("Ghi chú")
            },
            use_container_width=True
        )
        
        if st.button("Lưu thay đổi trận đấu", type="primary"):
            success, msg = save_editor_changes(db, "matches", "matches_editor", df_matches)
            if success:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)
    else:
        st.info("Chưa có trận đấu nào được lưu trong khoảng thời gian này.")


# =========================================================
# TAB 2: BẢNG XẾP HẠNG (RANKINGS)
# =========================================================
with tab_rank:
    st.subheader("Bảng xếp hạng thành tích")
    stats_list = db.calculate_stats(months_to_query)
    if stats_list:
        df_stats = pd.DataFrame(stats_list)
        df_stats["total_points"] = df_stats["wins"] * rules["point_win"] + df_stats["draws"] * rules["point_draw"]
        df_stats = df_stats.sort_values(by=["total_points", "wins", "gf"], ascending=[False, False, False])
        df_stats.insert(0, "Hạng", range(1, len(df_stats) + 1))
        
        df_stats_display = df_stats.drop(columns=["gf", "ga"], errors="ignore")
        df_stats_display = df_stats_display.rename(columns={
            "name": "Tên hội viên",
            "matches": "Trận",
            "wins": "Thắng",
            "draws": "Hòa",
            "losses": "Thua",
            "match_money": "Tiền trận",
            "bet_money": "Tiền độ",
            "money": "Tổng tiền",
            "total_points": "Điểm thành tích"
        })
        st.dataframe(
            df_stats_display.set_index("Hạng"),
            column_config={
                "Trận": st.column_config.NumberColumn("Trận", format="%d"),
                "Thắng": st.column_config.NumberColumn("Thắng", format="%d"),
                "Hòa": st.column_config.NumberColumn("Hòa", format="%d"),
                "Thua": st.column_config.NumberColumn("Thua", format="%d"),
                "Tiền trận": st.column_config.NumberColumn("Tiền trận", format="%,.0f"),
                "Tiền độ": st.column_config.NumberColumn("Tiền độ", format="%,.0f"),
                "Tổng tiền": st.column_config.NumberColumn("Tổng tiền", format="%,.0f"),
                "Điểm thành tích": st.column_config.NumberColumn("Điểm thành tích", format="%d")
            },
            use_container_width=True
        )
    else:
        st.info("Chưa có dữ liệu xếp hạng.")


# =========================================================
# TAB 3: BẢNG TIỀN TRẬN (FINES DETAIL)
# =========================================================
with tab_money:
    st.subheader("Chi tiết tiền trận & Tiền độ hội viên")
    money_detail = db.get_period_match_money_detail_map(months_to_query)
    if money_detail:
        money_rows = []
        for name, data in money_detail.items():
            money_rows.append({
                "Tên hội viên": name,
                "Tiền trận": data["match_money"],
                "Tiền độ": data["bet_money"],
                "Tổng cộng": data["total"]
            })
        df_money = pd.DataFrame(money_rows)
        df_money = df_money.sort_values(by="Tổng cộng", ascending=False)
        st.dataframe(
            df_money.set_index("Tên hội viên"),
            column_config={
                "Tiền trận": st.column_config.NumberColumn("Tiền trận", format="%,.0f"),
                "Tiền độ": st.column_config.NumberColumn("Tiền độ", format="%,.0f"),
                "Tổng cộng": st.column_config.NumberColumn("Tổng cộng", format="%,.0f")
            },
            use_container_width=True
        )
    else:
        st.info("Không có phát sinh tiền phạt hoặc tiền độ trong thời gian này.")


# =========================================================
# TAB 4: THỐNG KÊ TRẬN (STATS OVERVIEW)
# =========================================================
with tab_stats:
    st.subheader("Thống kê số liệu giải đấu")
    matches_raw = db.fetch_matches(months_to_query)
    if matches_raw:
        # Số liệu tóm tắt
        total_matches = len(matches_raw)
        total_collected = sum(
            get_match_collect_total(m[3], m[4], m[5], m[6], m[7], m[8], m[9], m[10], rules)
            for m in matches_raw
        )
        total_bets = sum(int(m[10] or 0) for m in matches_raw)
        
        c1, c2, c3 = st.columns(3)
        c1.metric("Tổng số trận đấu", f"{total_matches} trận")
        c2.metric("Tổng tiền thu từ trận", f"{total_collected:,.0f}")
        c3.metric("Tổng tiền độ phát sinh", f"{total_bets:,.0f}")
        
        st.write("---")
        st.subheader("Danh sách tất cả trận đấu trong kỳ")
        
        # Hiển thị list tĩnh đẹp đẽ để xem
        df_view = pd.DataFrame(matches_raw, columns=[
            "ID", "Ngày", "Trận số", "A1", "A2", "B1", "B2", "Điểm A", "Điểm B", "Tiền trận", "Tiền độ", "Ghi chú"
        ])
        df_view["Đội A"] = df_view.apply(lambda r: f"{r['A1']} + {r['A2']}" if r['A2'] else r['A1'], axis=1)
        df_view["Đội B"] = df_view.apply(lambda r: f"{r['B1']} + {r['B2']}" if r['B2'] else r['B1'], axis=1)
        df_view["Tỷ số"] = df_view.apply(lambda r: f"{r['Điểm A']} - {r['Điểm B']}", axis=1)
        
        df_display_list = df_view[["ID", "Ngày", "Trận số", "Đội A", "Tỷ số", "Đội B", "Tiền trận", "Tiền độ", "Ghi chú"]]
        st.dataframe(
            df_display_list.set_index("ID"),
            column_config={
                "Trận số": st.column_config.NumberColumn("Trận số", format="%d"),
                "Tiền trận": st.column_config.NumberColumn("Tiền trận", format="%,.0f"),
                "Tiền độ": st.column_config.NumberColumn("Tiền độ", format="%,.0f")
            },
            use_container_width=True
        )
    else:
        st.info("Chưa có trận đấu nào trong thời gian này.")


# =========================================================
# TAB 5: THU CHI THÁNG (MONTHLY FINANCE & SETTLEMENT)
# =========================================================
with tab_finance:
    st.subheader("Thẻ Tổng Hợp Tài Chính Quỹ")
    
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    col_f1.metric("Số dư quỹ lũy kế cuối kỳ", f"{finance_summary['fund_balance']:,.0f}")
    col_f2.metric("Tổng thu quỹ trong kỳ", f"{finance_summary['total_income_fund']:,.0f}")
    col_f3.metric("Tổng chi tiêu quỹ", f"{finance_summary['total_expense']:,.0f}")
    col_f4.metric("Thặng dư tích lũy kỳ", f"{finance_summary['fund_this_month']:,.0f}")
    
    st.write("---")
    
    # Grid điều chỉnh thu chi
    subtab_input, subtab_set = st.tabs(["💸 Quản lý Thu - Chi", "📊 Bảng Quyết Toán Tài Chính"])
    
    with subtab_input:
        # =========================================================
        # 1. PHẦN CHI (EXPENSES)
        # =========================================================
        st.markdown("### 1. Phần chi")
        
        # Hàng ngang nhập liệu cho khoản chi
        col_e1, col_e2, col_e3, col_e4, col_e5, col_e6 = st.columns([1.2, 1.5, 1.5, 1.5, 1.5, 1.2])
        with col_e1:
            new_exp_date = st.date_input("Ngày chi", value=datetime.now(), key="new_exp_date_input")
        with col_e2:
            new_exp_type = st.selectbox("Loại chi", options=EXPENSE_TYPES, index=0, key="new_exp_type_select")
        with col_e3:
            new_exp_amount = st.number_input("Số tiền chi", min_value=0, step=5000, value=0, key="new_exp_amount_input")
        with col_e4:
            new_exp_payer = st.selectbox("Người chi", options=["Quỹ"] + players_list, index=0, key="new_exp_payer_select")
        with col_e5:
            with st.popover("👥 Người tham gia", use_container_width=True):
                new_exp_parts = st.multiselect("Chọn hội viên:", options=["All"] + players_list, default=["All"], key="new_exp_parts_select")
                new_exp_guest_text = st.text_input("Người ngoài (dấu phẩy):", placeholder="Anh Nam, Anh Hải...", key="new_exp_guest_input")
                new_exp_note = st.text_input("Ghi chú chi:", placeholder="Ghi chú chi...", key="new_exp_note_input")
        with col_e6:
            st.write("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            add_exp_clicked = st.button("Thêm chi", type="primary", use_container_width=True, key="new_exp_add_btn")
            
        if add_exp_clicked:
            if new_exp_amount <= 0:
                st.error("Vui lòng nhập số tiền chi lớn hơn 0!")
            else:
                participants = list(new_exp_parts)
                guest_names = [x.strip() for x in new_exp_guest_text.replace(";", ",").split(",") if x.strip()]
                if "All" in participants:
                    participants = ["All"] + guest_names
                else:
                    participants = participants + guest_names
                participants = list(dict.fromkeys([p for p in participants if str(p).strip()]))
                if not participants:
                    participants = ["All"]
                    
                expense_date_str = new_exp_date.strftime("%Y-%m-%d")
                month_str = expense_date_str[:7]
                
                cur = db.conn.cursor()
                try:
                    cur.execute(
                        """
                        INSERT INTO finance_expenses (month, expense_date, expense_type, amount, participants, payer, note)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (month_str, expense_date_str, new_exp_type, new_exp_amount, json.dumps(participants, ensure_ascii=False), new_exp_payer, new_exp_note)
                    )
                    db.conn.commit()
                    st.success("Đã thêm khoản chi thành công!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")

        # Hiển thị bảng chi
        if not expenses_list:
            st.info("💡 Chưa có khoản chi nào trong kỳ này. Nhập các trường trên và bấm 'Thêm chi' để bắt đầu.")
            df_expenses = pd.DataFrame(columns=["ID", "Ngày chi", "Loại chi phí", "Số tiền", "Người tham gia", "Người trả", "Ghi chú"])
        else:
            df_expenses = pd.DataFrame(expenses_list, columns=["ID", "Ngày chi", "Loại chi phí", "Số tiền", "Người tham gia", "Người trả", "Ghi chú"])
            df_expenses["Người tham gia"] = df_expenses["Người tham gia"].apply(lambda x: ", ".join(load_json_list(x)) if "All" not in load_json_list(x) else "All")
            
        edited_exp_df = st.data_editor(
            df_expenses,
            num_rows="dynamic",
            key="expenses_editor",
            column_config={
                "ID": st.column_config.NumberColumn("ID", disabled=True, format="%d"),
                "Ngày chi": st.column_config.TextColumn("Ngày chi"),
                "Loại chi phí": st.column_config.SelectboxColumn("Loại chi phí", options=EXPENSE_TYPES),
                "Số tiền": st.column_config.NumberColumn("Số tiền", format="%,.0f"),
                "Người tham gia": st.column_config.TextColumn("Người tham gia"),
                "Người trả": st.column_config.TextColumn("Người trả"),
                "Ghi chú": st.column_config.TextColumn("Ghi chú")
            },
            use_container_width=True
        )
        
        if st.button("Lưu thay đổi khoản chi", type="primary", key="save_exp_btn"):
            success, msg = save_editor_changes(db, "finance_expenses", "expenses_editor", df_expenses)
            if success:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)
                
        st.write("---")

        # =========================================================
        # 2. PHẦN THU (INCOMES)
        # =========================================================
        st.markdown("### 2. Phần thu")
        
        # Hàng ngang nhập liệu cho khoản thu
        col_i1, col_i2, col_i3, col_i4, col_i5, col_i6 = st.columns([1.2, 1.5, 1.5, 1.5, 1.5, 1.2])
        with col_i1:
            new_inc_date = st.date_input("Ngày thu", value=datetime.now(), key="new_inc_date_input")
        with col_i2:
            new_inc_type = st.selectbox("Loại thu", options=DEFAULT_INCOME_TYPES, index=0, key="new_inc_type_select")
        with col_i3:
            new_inc_amount = st.number_input("Số tiền thu", min_value=0, step=5000, value=0, key="new_inc_amount_input")
        with col_i4:
            new_inc_collector = st.selectbox("Người thu", options=["Quỹ"] + players_list, index=0, key="new_inc_collector_select")
        with col_i5:
            new_inc_note = st.text_input("Ghi chú thu", placeholder="Nhập ghi chú...", key="new_inc_note_input")
        with col_i6:
            st.write("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            add_inc_clicked = st.button("Thêm thu", type="primary", use_container_width=True, key="new_inc_add_btn")
            
        if add_inc_clicked:
            if new_inc_amount <= 0:
                st.error("Vui lòng nhập số tiền thu lớn hơn 0!")
            else:
                income_date_str = new_inc_date.strftime("%Y-%m-%d")
                month_str = income_date_str[:7]
                
                cur = db.conn.cursor()
                try:
                    cur.execute(
                        """
                        INSERT INTO finance_incomes (month, income_date, income_type, amount, collector, note)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (month_str, income_date_str, new_inc_type, new_inc_amount, new_inc_collector, new_inc_note)
                    )
                    db.conn.commit()
                    st.success("Đã thêm khoản thu thành công!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Lỗi: {e}")

        # Hiển thị bảng thu
        target_month = months_to_query[-1]
        rounded_collector = db.get_auto_income_collector(AUTO_ROUNDED_MEMBER_INCOME, target_month)
        external_collector = db.get_auto_income_collector(AUTO_EXTERNAL_INCOME, target_month)
        
        auto_rows = [
            ("AUTO", target_month, AUTO_ROUNDED_MEMBER_INCOME, int(round(finance_summary.get("total_rounded_fee", 0))), rounded_collector, ""),
            ("AUTO", target_month, AUTO_EXTERNAL_INCOME, int(round(finance_summary.get("external_income_total", 0))), external_collector, "")
        ]
        df_incomes_auto = pd.DataFrame(auto_rows, columns=["ID", "Ngày thu", "Loại khoản thu", "Số tiền", "Người thu", "Ghi chú"])
        
        if not incomes_list:
            df_incomes_manual = pd.DataFrame(columns=["ID", "Ngày thu", "Loại khoản thu", "Số tiền", "Người thu", "Ghi chú"])
        else:
            df_incomes_manual = pd.DataFrame(incomes_list, columns=["ID", "Ngày thu", "Loại khoản thu", "Số tiền", "Người thu", "Ghi chú"])
            
        df_incomes = pd.concat([df_incomes_auto, df_incomes_manual], ignore_index=True)
            
        edited_inc_df = st.data_editor(
            df_incomes,
            num_rows="dynamic",
            key="incomes_editor",
            column_config={
                "ID": st.column_config.TextColumn("ID", disabled=True),
                "Ngày thu": st.column_config.TextColumn("Ngày thu"),
                "Loại khoản thu": st.column_config.SelectboxColumn("Loại khoản thu", options=DEFAULT_INCOME_TYPES + list(AUTO_INCOME_TYPES)),
                "Số tiền": st.column_config.NumberColumn("Số tiền", format="%,.0f"),
                "Người thu": st.column_config.TextColumn("Người thu"),
                "Ghi chú": st.column_config.TextColumn("Ghi chú")
            },
            use_container_width=True
        )
        
        if st.button("Lưu thay đổi khoản thu", type="primary", key="save_inc_btn"):
            success, msg = save_editor_changes(db, "finance_incomes", "incomes_editor", df_incomes)
            if success:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    with subtab_set:
        st.subheader("Báo cáo quyết toán chi tiết")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            excel_data = export_excel_bytes(db, months_to_query)
            st.download_button(
                label="📥 Tải Báo cáo Excel đầy đủ",
                data=excel_data,
                file_name=f"Bao_cao_Tennis_{selected_period}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_dl2:
            pdf_data = export_pdf_bytes(db, months_to_query)
            st.download_button(
                label="📥 Tải Báo cáo PDF",
                data=pdf_data,
                file_name=f"Bao_cao_Tennis_{selected_period}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
            
        st.write("---")
        st.markdown("#### 1. Quyết toán Hội viên")
        if member_finance_rows:
            df_mem_fin = pd.DataFrame(member_finance_rows)
            df_mem_fin_display = df_mem_fin.rename(columns={
                "name": "Hội viên", "weekly": "Buổi/Tuần", "court": "Tiền sân", "ball_picker": "Bóng + Nhặt",
                "meal": "Ăn uống", "other_charge": "Chi khác", "match_money": "Tiền phạt", "bet_money": "Tiền độ",
                "calc_fee": "Phí thực tính", "rounded_fee": "Phí làm tròn", "paid_by_member": "Đã chi hộ",
                "collected_by_member": "Đã thu hộ", "net_payable": "Cần thu (Thu tháng)"
            })
            cols_to_show = ["Hội viên", "Buổi/Tuần", "Tiền sân", "Bóng + Nhặt", "Ăn uống", "Chi khác", "Tiền phạt", "Tiền độ", "Phí thực tính", "Phí làm tròn", "Đã chi hộ", "Đã thu hộ", "Cần thu (Thu tháng)"]
            st.dataframe(
                df_mem_fin_display[cols_to_show].set_index("Hội viên"),
                column_config={
                    "Buổi/Tuần": st.column_config.NumberColumn("Buổi/Tuần", format="%d"),
                    "Tiền sân": st.column_config.NumberColumn("Tiền sân", format="%,.0f"),
                    "Bóng + Nhặt": st.column_config.NumberColumn("Bóng + Nhặt", format="%,.0f"),
                    "Ăn uống": st.column_config.NumberColumn("Ăn uống", format="%,.0f"),
                    "Chi khác": st.column_config.NumberColumn("Chi khác", format="%,.0f"),
                    "Tiền phạt": st.column_config.NumberColumn("Tiền phạt", format="%,.0f"),
                    "Tiền độ": st.column_config.NumberColumn("Tiền độ", format="%,.0f"),
                    "Phí thực tính": st.column_config.NumberColumn("Phí thực tính", format="%,.0f"),
                    "Phí làm tròn": st.column_config.NumberColumn("Phí làm tròn", format="%,.0f"),
                    "Đã chi hộ": st.column_config.NumberColumn("Đã chi hộ", format="%,.0f"),
                    "Đã thu hộ": st.column_config.NumberColumn("Đã thu hộ", format="%,.0f"),
                    "Cần thu (Thu tháng)": st.column_config.NumberColumn("Cần thu (Thu tháng)", format="%,.0f")
                },
                use_container_width=True
            )
            
        st.write("---")
        st.markdown("#### 2. Quyết toán Người ngoài / Khách vãng lai")
        ext_rows = finance_summary["external_rows"]
        if ext_rows:
            df_ext = pd.DataFrame(ext_rows)
            df_ext_display = df_ext.rename(columns={
                "name": "Họ và tên", "meal": "Tiền ăn uống", "other": "Khoản chi khác",
                "paid_by_external": "Đã chi hộ", "collected_by_external": "Đã thu hộ", "net_payable": "Cần thu/trả lại"
            })
            st.dataframe(
                df_ext_display.set_index("Họ và tên"),
                column_config={
                    "Tiền ăn uống": st.column_config.NumberColumn("Tiền ăn uống", format="%,.0f"),
                    "Khoản chi khác": st.column_config.NumberColumn("Khoản chi khác", format="%,.0f"),
                    "Đã chi hộ": st.column_config.NumberColumn("Đã chi hộ", format="%,.0f"),
                    "Đã thu hộ": st.column_config.NumberColumn("Đã thu hộ", format="%,.0f"),
                    "Cần thu/trả lại": st.column_config.NumberColumn("Cần thu/trả lại", format="%,.0f")
                },
                use_container_width=True
            )
        else:
            st.info("Không phát sinh quyết toán của người ngoài.")


# =========================================================
# TAB 6: DỮ LIỆU (DATA & CONFIGURATION)
# =========================================================
with tab_data:
    subtab_members, subtab_rules = st.tabs(["👥 Danh sách Hội Viên", "⚙️ Luật & Thủ Quỹ"])
    
    with subtab_members:
        st.subheader("Danh sách hội viên đang hoạt động")
        cur = db.conn.cursor()
        raw_players = cur.execute("SELECT id, name, active, weekly_sessions FROM players ORDER BY name").fetchall()
        
        df_players = pd.DataFrame(raw_players, columns=["ID", "Họ và tên", "Hoạt động", "Số buổi/Tuần"])
        df_players["Hoạt động"] = df_players["Hoạt động"].apply(lambda x: True if x == 1 else False)
        
        edited_players_df = st.data_editor(
            df_players,
            num_rows="dynamic",
            key="players_editor",
            column_config={
                "ID": st.column_config.NumberColumn("ID", disabled=True),
                "Họ và tên": st.column_config.TextColumn("Họ và tên"),
                "Hoạt động": st.column_config.CheckboxColumn("Hoạt động"),
                "Số buổi/Tuần": st.column_config.SelectboxColumn("Số buổi/Tuần", options=[1, 2, 3])
            },
            use_container_width=True
        )
        
        if st.button("Lưu thay đổi hội viên", type="primary", key="save_players_btn"):
            success, msg = save_editor_changes(db, "players", "players_editor", df_players)
            if success:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    with subtab_rules:
        st.subheader("Cấu hình mức phạt & Tính điểm")
        
        rule_lose = st.number_input("Tiền phạt khi thua trận:", value=rules["fine_lose"], step=5000)
        rule_draw = st.number_input("Tiền phạt khi hòa trận:", value=rules["fine_draw"], step=5000)
        rule_zero = st.number_input("Tiền phạt khi thua trắng 0 bàn:", value=rules["fine_lose_zero"], step=5000)
        
        rule_pt_win = st.number_input("Điểm cộng khi Thắng:", value=rules["point_win"], step=1)
        rule_pt_draw = st.number_input("Điểm cộng khi Hòa:", value=rules["point_draw"], step=1)
        rule_pt_loss = st.number_input("Điểm cộng khi Thua:", value=rules["point_loss"], step=1)
        
        st.write("---")
        st.subheader("Thiết lập thủ quỹ tự động thu")
        players_names = db.get_player_names()
        
        # Dùng tháng cuối cùng trong kỳ truy vấn để lưu/đọc thiết lập thủ quỹ
        target_month = months_to_query[-1]
        
        auto_collector_rounded = st.selectbox(
            "Người thu hộ tiền hội viên làm tròn:",
            options=["Quỹ"] + players_names,
            index=(["Quỹ"] + players_names).index(db.get_auto_income_collector(AUTO_ROUNDED_MEMBER_INCOME, target_month))
        )
        
        auto_collector_external = st.selectbox(
            "Người thu hộ tiền người ngoài vãng lai:",
            options=["Quỹ"] + players_names,
            index=(["Quỹ"] + players_names).index(db.get_auto_income_collector(AUTO_EXTERNAL_INCOME, target_month))
        )

        if st.button("Lưu cấu hình", key="save_rules_btn"):
            new_rules = {
                "fine_lose": rule_lose,
                "fine_draw": rule_draw,
                "fine_lose_zero": rule_zero,
                "point_win": rule_pt_win,
                "point_draw": rule_pt_draw,
                "point_loss": rule_pt_loss,
            }
            db.save_rules(new_rules)
            
            db.conn.execute(
                "INSERT OR REPLACE INTO settings(key, value) VALUES (?, ?)",
                (f"auto_income_collector::{target_month}::{AUTO_ROUNDED_MEMBER_INCOME}", auto_collector_rounded)
            )
            db.conn.execute(
                "INSERT OR REPLACE INTO settings(key, value) VALUES (?, ?)",
                (f"auto_income_collector::{target_month}::{AUTO_EXTERNAL_INCOME}", auto_collector_external)
            )
            db.conn.commit()
            
            st.success("Đã lưu các thiết lập cấu hình mới!")
            st.rerun()

# Đóng kết nối khi ứng dụng chạy xong
db.conn.close()
